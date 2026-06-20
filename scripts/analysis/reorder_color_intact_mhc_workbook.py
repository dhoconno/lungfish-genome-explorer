#!/usr/bin/env python3
"""Build a chromosome-ordered, color-coded MHC haplotype review workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import platform
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.views import Pane


LOCUS_ORDER = ["MHC-A", "MHC-B", "MHC-DR", "MHC-DQ", "MHC-DP"]
LOCUS_RANK = {name: i for i, name in enumerate(LOCUS_ORDER)}

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SUBHEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_SECTION = PatternFill("solid", fgColor="E2F0D9")
FILL_H1 = PatternFill("solid", fgColor="D9EAF7")
FILL_H2 = PatternFill("solid", fgColor="FCE4D6")
FILL_BOTH = PatternFill("solid", fgColor="C6EFCE")
FILL_OFFCALL = PatternFill("solid", fgColor="E7E6E6")
FILL_INVALID = PatternFill("solid", fgColor="F4CCCC")
FILL_WARNING = PatternFill("solid", fgColor="FFF2CC")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

THIN_GREY = Side(style="thin", color="D9E1F2")
MEDIUM_BLUE = Side(style="medium", color="1F4E79")
GRID = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def file_record(path: Path, role: str) -> dict:
    return {
        "path": str(path),
        "role": role,
        "sha256": sha256_file(path),
        "sizeBytes": path.stat().st_size,
    }


def parse_pipe_fields(label: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    if not isinstance(label, str):
        return fields
    for part in label.split("|")[1:]:
        if "=" in part:
            key, value = part.split("=", 1)
            fields[key] = value
    return fields


def mcm_index(label: str) -> int:
    if not isinstance(label, str):
        return 10**9
    match = re.search(r"MCM_MHC_MiSeq_(\d+)", label)
    return int(match.group(1)) if match else 10**9


def source_locus_key(label: str) -> str:
    fields = parse_pipe_fields(label)
    return fields.get("source_loci", "")


def hap_group(label: str) -> str:
    fields = parse_pipe_fields(label)
    return fields.get("haplotype_groups", "Other")


def haplotypes(label: str) -> set[str]:
    raw = parse_pipe_fields(label).get("haplotypes", "")
    return {item.strip() for item in raw.split(",") if item.strip()}


def row_sort_key(row_values: list) -> tuple:
    label = row_values[0]
    group = hap_group(label)
    rank = LOCUS_RANK.get(group, len(LOCUS_ORDER))
    return (rank, source_locus_key(label), mcm_index(label), str(label))


def copy_matrix(ws) -> list[list]:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def autosize_columns(ws, min_width: int = 8, max_width: int = 56) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[letter].width = 58
            continue
        if col_idx in (2, 3):
            ws.column_dimensions[letter].width = 12
            continue
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def style_main_sheet(ws, sample_columns: list[int]) -> None:
    ws.freeze_panes = "D18"
    ws.sheet_view.pane = Pane(xSplit=3, ySplit=17, topLeftCell="D18", activePane="bottomRight", state="frozen")
    ws.auto_filter.ref = f"A17:{get_column_letter(ws.max_column)}{ws.max_row}"

    for row in ws.iter_rows():
        for cell in row:
            cell.border = GRID
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

    for row_idx in (1, 2):
        for cell in ws[row_idx]:
            cell.fill = FILL_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(3, 6):
        ws.cell(row_idx, 1).fill = FILL_SUBHEADER
        ws.cell(row_idx, 1).font = Font(bold=True)

    for row_idx in range(6, 16):
        fill = FILL_H1 if row_idx % 2 == 0 else FILL_H2
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).fill = fill
        ws.cell(row_idx, 1).font = Font(bold=True)

    for cell in ws[16]:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx in range(17, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        if isinstance(label, str) and label.endswith("allele evidence"):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = FILL_SECTION
                cell.font = Font(bold=True)
                cell.border = Border(top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)

    for col_idx in sample_columns:
        ws.cell(1, col_idx).alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
        ws.cell(2, col_idx).alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
    ws.row_dimensions[1].height = 78
    ws.row_dimensions[2].height = 78
    autosize_columns(ws)
    for col_idx in sample_columns:
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
        ws.cell(16, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[16].height = 42


def build_workbook(input_xlsx: Path, output_xlsx: Path, genotype_csv: Path, annotations: Path) -> dict:
    source = load_workbook(input_xlsx, data_only=False)
    main_in = source[source.sheetnames[0]]
    evidence_in = source["Intact MHC Call Evidence"]

    main_rows = copy_matrix(main_in)
    header_rows = main_rows[:16]
    allele_rows = [row for row in main_rows[17:] if isinstance(row[0], str) and row[0].startswith("MCM_MHC_MiSeq_")]
    sorted_allele_rows = sorted(allele_rows, key=row_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordered MHC calls"

    for row in header_rows:
        ws.append(row)

    row_to_group: dict[int, str] = {}
    for group in LOCUS_ORDER:
        group_rows = [row for row in sorted_allele_rows if hap_group(row[0]) == group]
        if not group_rows:
            continue
        ws.append([f"{group} allele evidence"] + [None] * (main_in.max_column - 1))
        for row in group_rows:
            ws.append(row)
            row_to_group[ws.max_row] = group

    other_rows = [row for row in sorted_allele_rows if hap_group(row[0]) not in LOCUS_RANK]
    if other_rows:
        ws.append(["Other allele evidence"] + [None] * (main_in.max_column - 1))
        for row in other_rows:
            ws.append(row)
            row_to_group[ws.max_row] = "Other"

    sample_by_col = {col_idx: ws.cell(1, col_idx).value for col_idx in range(4, ws.max_column + 1)}
    sample_columns = list(sample_by_col)

    hap_row_by_group = {}
    for row_idx in range(6, 16):
        label = ws.cell(row_idx, 1).value
        if not isinstance(label, str):
            continue
        match = re.match(r"(MHC-[A-Z]+) Haplotype ([12])", label)
        if match:
            hap_row_by_group[(match.group(1), match.group(2))] = row_idx

    for col_idx, sample in sample_by_col.items():
        if sample == "LF2840":
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row_idx, col_idx).fill = FILL_INVALID

    support_counts = defaultdict(int)
    for row_idx, group in row_to_group.items():
        label = ws.cell(row_idx, 1).value
        row_haps = haplotypes(label)
        for col_idx, sample in sample_by_col.items():
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value in (None, "", 0, "0"):
                continue
            if sample == "LF2840":
                cell.fill = FILL_INVALID
                support_counts["invalid_sample_cells"] += 1
                continue
            h1_row = hap_row_by_group.get((group, "1"))
            h2_row = hap_row_by_group.get((group, "2"))
            h1 = str(ws.cell(h1_row, col_idx).value).strip() if h1_row and ws.cell(h1_row, col_idx).value else ""
            h2 = str(ws.cell(h2_row, col_idx).value).strip() if h2_row and ws.cell(h2_row, col_idx).value else ""
            supports_h1 = h1 in row_haps if h1 else False
            supports_h2 = h2 in row_haps if h2 else False
            if supports_h1 and supports_h2:
                cell.fill = FILL_BOTH
                support_counts["supports_both"] += 1
            elif supports_h1:
                cell.fill = FILL_H1
                support_counts["supports_haplotype_1"] += 1
            elif supports_h2:
                cell.fill = FILL_H2
                support_counts["supports_haplotype_2"] += 1
            else:
                cell.fill = FILL_OFFCALL
                support_counts["observed_off_call"] += 1

    style_main_sheet(ws, sample_columns)

    evidence_ws = wb.create_sheet("Intact MHC Call Evidence")
    evidence_rows = copy_matrix(evidence_in)
    evidence_ws.append(evidence_rows[0])
    locus_col = evidence_rows[0].index("locus")
    sample_col = evidence_rows[0].index("sample")
    for row in sorted(evidence_rows[1:], key=lambda r: (str(r[sample_col]), LOCUS_RANK.get(r[locus_col], 99))):
        evidence_ws.append(row)
    evidence_ws.freeze_panes = "A2"
    evidence_ws.auto_filter.ref = f"A1:H{evidence_ws.max_row}"
    for cell in evidence_ws[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
    for row in evidence_ws.iter_rows():
        for cell in row:
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_columns(evidence_ws, max_width=72)

    legend = wb.create_sheet("Legend")
    legend_rows = [
        ["Item", "Meaning"],
        ["Row order", "Allele evidence rows are grouped MHC-A, MHC-B, MHC-DR, MHC-DQ, MHC-DP, then any other group; within each group rows are sorted by source_loci and MCM_MHC_MiSeq index."],
        ["Blue genotype cell", "Observed allele call carries the haplotype tag assigned as Haplotype 1 for that sample and locus block."],
        ["Orange genotype cell", "Observed allele call carries the haplotype tag assigned as Haplotype 2 for that sample and locus block."],
        ["Green genotype cell", "Observed allele call carries both assigned haplotype tags, or the sample is homozygous for that block."],
        ["Grey genotype cell", "Observed allele call does not carry either assigned haplotype tag for that sample and locus block."],
        ["Red LF2840 column", "LF2840 is flagged as invalid for this pre-fix demultiplexing artifact and should not be interpreted."],
        ["MHC-DP note", "No MHC-DB group was present in the reference/genotype labels; this workbook uses the MHC-DP block."],
    ]
    for row in legend_rows:
        legend.append(row)
    for cell in legend[1]:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
    for row in legend.iter_rows():
        for cell in row:
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    legend.column_dimensions["A"].width = 24
    legend.column_dimensions["B"].width = 112
    legend["A3"].fill = FILL_H1
    legend["A4"].fill = FILL_H2
    legend["A5"].fill = FILL_BOTH
    legend["A6"].fill = FILL_OFFCALL
    legend["A7"].fill = FILL_INVALID

    if legend.max_row >= 2:
        tab = Table(displayName="LegendTable", ref=f"A1:B{legend.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        legend.add_table(tab)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    return {
        "alleleRows": len(sorted_allele_rows),
        "groupedRows": {group: sum(1 for row in sorted_allele_rows if hap_group(row[0]) == group) for group in LOCUS_ORDER},
        "otherRows": len(other_rows),
        "sampleCount": len(sample_columns),
        "supportColorCounts": dict(sorted(support_counts.items())),
        "inputs": [input_xlsx, genotype_csv, annotations],
    }


def write_provenance(
    output_xlsx: Path,
    provenance_path: Path,
    input_xlsx: Path,
    genotype_csv: Path,
    annotations: Path,
    started: float,
    result: dict,
    exit_status: int,
    stderr: str = "",
) -> None:
    now = time.time()
    command = " ".join([sys.executable, *sys.argv])
    payload = {
        "schemaVersion": 1,
        "toolName": "codex-python-openpyxl ordered-color-coded MHC workbook builder",
        "toolVersion": "local",
        "workflowName": "chromosome-ordered color-coded MHC haplotype workbook export",
        "argv": [sys.executable, *sys.argv],
        "command": command,
        "startedAt": datetime.fromtimestamp(started, timezone.utc).isoformat().replace("+00:00", "Z"),
        "completedAt": datetime.fromtimestamp(now, timezone.utc).isoformat().replace("+00:00", "Z"),
        "wallTimeSeconds": round(now - started, 3),
        "exitStatus": exit_status,
        "stderr": stderr,
        "runtimeIdentity": {
            "python": platform.python_version(),
            "platform": platform.platform(),
            "executable": sys.executable,
        },
        "options": {
            "locusOrder": LOCUS_ORDER,
            "rowSort": "group by haplotype_groups, then source_loci, then MCM_MHC_MiSeq numeric index",
            "colorCoding": {
                "blue": "allele supports assigned haplotype 1",
                "orange": "allele supports assigned haplotype 2",
                "green": "allele supports both assigned haplotypes",
                "grey": "observed allele does not support either assigned haplotype",
                "red": "LF2840 invalid pre-fix demultiplexing artifact",
            },
            "dpNote": "No MHC-DB group was found; MHC-DP is used for the final class-II block.",
            "resolvedDefaults": {
                "firstSampleColumn": "D",
                "headerRowsPreserved": 16,
                "inputDataRowsStart": 18,
                "outputSheets": ["Ordered MHC calls", "Intact MHC Call Evidence", "Legend"],
            },
        },
        "summary": {k: v for k, v in result.items() if k != "inputs"},
        "inputs": [
            file_record(input_xlsx, "input"),
            file_record(genotype_csv, "input"),
            file_record(annotations, "input"),
        ],
        "outputs": [file_record(output_xlsx, "output")],
    }
    provenance_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--genotype-csv", required=True, type=Path)
    parser.add_argument("--annotations", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    args = parser.parse_args()

    started = time.time()
    stderr = ""
    exit_status = 1
    try:
        result = build_workbook(args.input_xlsx, args.output_xlsx, args.genotype_csv, args.annotations)
        exit_status = 0
        write_provenance(
            args.output_xlsx,
            args.output_xlsx.with_suffix(args.output_xlsx.suffix + ".lungfish-provenance.json"),
            args.input_xlsx,
            args.genotype_csv,
            args.annotations,
            started,
            result,
            exit_status,
            stderr,
        )
        print(json.dumps(result, indent=2, sort_keys=True, default=str))
        return 0
    except Exception as exc:  # pragma: no cover - command-line provenance path
        stderr = f"{type(exc).__name__}: {exc}"
        if args.output_xlsx.exists():
            try:
                write_provenance(
                    args.output_xlsx,
                    args.output_xlsx.with_suffix(args.output_xlsx.suffix + ".lungfish-provenance.json"),
                    args.input_xlsx,
                    args.genotype_csv,
                    args.annotations,
                    started,
                    {},
                    exit_status,
                    stderr,
                )
            except Exception:
                pass
        print(stderr, file=sys.stderr)
        return exit_status


if __name__ == "__main__":
    raise SystemExit(main())
