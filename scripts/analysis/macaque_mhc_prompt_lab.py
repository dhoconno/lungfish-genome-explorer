#!/usr/bin/env python3
"""Evaluation-only prompt lab for de novo macaque MHC haplotyping."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import re
import shlex
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict, deque
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TOOL_NAME = "macaque-mhc-prompt-lab"
TOOL_VERSION = "2026-06-20.1"
PROMPT_VERSION = "generalist_macaque_mhc_haplotyping_v1"
REPORT_LOCI = ["MHC-A", "MHC-B", "MHC-DRB", "MHC-DQA", "MHC-DQB", "MHC-DPA", "MHC-DPB"]
FULL_RESULT_SHEETS = ["Full Sequencing Results 1", "Full Sequencing Results 2"]
DUPLICATE_SAMPLE_POLICY = "FULL_RESULT_SHEETS are ordered older-to-newer; later duplicate gs_id entries supersede earlier sample, truth, and observations."
DEFAULT_SNPRC_WORKBOOK = Path("/Users/dho/Downloads/30783_SNPRC22_MHC_Genotype_Report_31Dec24.xlsx")
DEFAULT_PROMPT = Path("scripts/analysis/prompts/generalist_macaque_mhc_haplotyping_v1.md")
DEFAULT_OUTPUT_ROOT = Path("outputs/macaque-mhc-prompt-lab")
DEFAULT_OPENAI_MODEL = "gpt-5.5"
DEFAULT_OPENAI_MAX_OUTPUT_TOKENS = 24000
DEFAULT_OPENAI_REASONING_EFFORT = "medium"
OPENAI_SYSTEM_PROMPT = "Follow the user prompt exactly and return only valid JSON."
SECRET_REDACTION_MARKER = "[REDACTED_SECRET]"
PROMPT_INPUT_PLACEHOLDER = "{{PROMPT_INPUT_JSON}}"
TRUTH_ROW_RE = re.compile(r"^MHC-(A|B|DRB|DQA|DQB|DPA|DPB)\s+Haplotype\s+([12])$", re.IGNORECASE)
GENOTYPE_PREFIX_RE = re.compile(r"^\d+_(Mamu-[A-Za-z0-9*]+)")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_record(path: Path, role: str) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "role": role,
        "sha256": sha256_file(path),
        "sizeBytes": path.stat().st_size,
    }


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def int_or_none(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def locus_from_genotype(genotype: str) -> tuple[str, str]:
    match = GENOTYPE_PREFIX_RE.match(genotype)
    source = match.group(1) if match else genotype.split("_", 1)[0]
    upper = source.upper()
    if upper.startswith("MAMU-A") or upper.startswith("MAMU-AG") or upper.startswith("MAMU-F") or upper.startswith("MAMU-G"):
        return ("MHC-A", source)
    if upper.startswith("MAMU-B") or upper.startswith("MAMU-I") or upper.startswith("MAMU-J") or upper.startswith("MAMU-K") or upper.startswith("MAMU-S") or upper.startswith("MAMU-V"):
        return ("MHC-B", source)
    if upper.startswith("MAMU-DRB"):
        return ("MHC-DRB", source)
    if upper.startswith("MAMU-DQA"):
        return ("MHC-DQA", source)
    if upper.startswith("MAMU-DQB"):
        return ("MHC-DQB", source)
    if upper.startswith("MAMU-DPA"):
        return ("MHC-DPA", source)
    if upper.startswith("MAMU-DPB"):
        return ("MHC-DPB", source)
    if upper.startswith("MAMU-E"):
        return ("context", source)
    return ("context", source)


def is_genotype_label(label: str) -> bool:
    return bool(re.match(r"^\d+_Mamu-", label))


def sample_columns(ws) -> list[int]:
    cols = []
    for col in range(4, ws.max_column + 1):
        first = clean(ws.cell(1, col).value)
        second = clean(ws.cell(2, col).value)
        if first or second:
            cols.append(col)
    return cols


def metadata_for_column(ws, col: int) -> dict[str, Any]:
    row1_label = clean(ws.cell(1, 1).value).lower().replace(" ", "_")
    row2_label = clean(ws.cell(2, 1).value).lower().replace(" ", "_")
    row1 = clean(ws.cell(1, col).value)
    row2 = clean(ws.cell(2, col).value)
    if row1_label in {"client_id", "clientid"} and row2_label in {"gs_id", "gsid"}:
        client_id, gs_id = row1, row2
    elif row1_label in {"gs_id", "gsid"} and row2_label in {"client_id", "clientid"}:
        gs_id, client_id = row1, row2
    else:
        client_id, gs_id = row1, row2
    mapped_reads = int_or_none(ws.cell(3, col).value)
    return {"client_id": client_id, "gs_id": gs_id or client_id, "mapped_reads": mapped_reads}


def row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def sample_column_indexes(row1: tuple[Any, ...], row2: tuple[Any, ...]) -> list[int]:
    cols = []
    for index in range(3, max(len(row1), len(row2))):
        first = clean(row_value(row1, index))
        second = clean(row_value(row2, index))
        if first or second:
            cols.append(index)
    return cols


def metadata_for_column_values(row1: tuple[Any, ...], row2: tuple[Any, ...], row3: tuple[Any, ...], index: int) -> dict[str, Any]:
    row1_label = clean(row_value(row1, 0)).lower().replace(" ", "_")
    row2_label = clean(row_value(row2, 0)).lower().replace(" ", "_")
    first_value = clean(row_value(row1, index))
    second_value = clean(row_value(row2, index))
    if row1_label in {"client_id", "clientid"} and row2_label in {"gs_id", "gsid"}:
        client_id, gs_id = first_value, second_value
    elif row1_label in {"gs_id", "gsid"} and row2_label in {"client_id", "clientid"}:
        gs_id, client_id = first_value, second_value
    else:
        client_id, gs_id = first_value, second_value
    mapped_reads = int_or_none(row_value(row3, index))
    return {"client_id": client_id, "gs_id": gs_id or client_id, "mapped_reads": mapped_reads}


def normalize_truth_locus(raw: str) -> str:
    return "MHC-" + raw.upper()


def validate_extracted_dataset(
    workbook_path: Path,
    samples: list[dict[str, Any]],
    truth: dict[str, dict[str, list[str]]],
    observations: list[dict[str, Any]],
) -> None:
    if not samples:
        raise ValueError(f"Workbook {workbook_path} yielded no samples")
    if not observations:
        raise ValueError(f"Workbook {workbook_path} yielded no genotype observations")
    if not truth:
        raise ValueError(f"Workbook {workbook_path} yielded no truth calls")
    sample_ids = {sample["gs_id"] for sample in samples}
    observation_sample_ids = {row["sample_id"] for row in observations}
    missing_truth = sorted((sample_ids | observation_sample_ids) - set(truth))
    if missing_truth:
        raise ValueError(f"Workbook {workbook_path} missing truth calls for samples: {', '.join(missing_truth)}")
    for sample_id in sorted(sample_ids | observation_sample_ids | set(truth)):
        calls = truth.get(sample_id, {})
        for locus in REPORT_LOCI:
            slots = calls.get(locus)
            if slots is None or len(slots) != 2 or any(not clean(slot) for slot in slots):
                raise ValueError(f"Workbook {workbook_path} has incomplete truth calls for {sample_id} {locus}")


def extract_sheet(ws) -> tuple[list[dict[str, Any]], dict[str, dict[str, list[str]]], list[dict[str, Any]]]:
    samples = []
    truth: dict[str, dict[str, list[str]]] = defaultdict(lambda: {locus: ["", ""] for locus in REPORT_LOCI})
    observations = []
    rows = ws.iter_rows(values_only=True)
    row1 = tuple(next(rows, ()))
    row2 = tuple(next(rows, ()))
    row3 = tuple(next(rows, ()))
    columns = sample_column_indexes(row1, row2)
    sample_meta_by_col = {col: metadata_for_column_values(row1, row2, row3, col) for col in columns}
    for meta in sample_meta_by_col.values():
        if meta["gs_id"]:
            samples.append({"gs_id": meta["gs_id"], "client_id": meta["client_id"], "sheet": ws.title, "mapped_reads": meta["mapped_reads"]})

    for row_number, row_values in enumerate(rows, start=4):
        row_values = tuple(row_values)
        label = clean(row_value(row_values, 0))
        if not label:
            continue
        truth_match = TRUTH_ROW_RE.match(label)
        if truth_match:
            locus = normalize_truth_locus(truth_match.group(1))
            slot = int(truth_match.group(2)) - 1
            for col, meta in sample_meta_by_col.items():
                sample_id = meta["gs_id"]
                if sample_id:
                    truth[sample_id][locus][slot] = clean(row_value(row_values, col))
            continue
        if not is_genotype_label(label):
            continue
        report_locus, source_locus = locus_from_genotype(label)
        if report_locus == "context":
            continue
        for col, meta in sample_meta_by_col.items():
            reads = int_or_none(row_value(row_values, col))
            if reads is None or reads <= 0 or not meta["gs_id"]:
                continue
            mapped = meta["mapped_reads"]
            observations.append({
                "sample_id": meta["gs_id"],
                "client_id": meta["client_id"],
                "report_locus": report_locus,
                "source_locus": source_locus,
                "genotype": label,
                "reads": reads,
                "sheet": ws.title,
                "row": row_number,
                "sample_mapped_reads": mapped,
                "read_fraction": round(reads / mapped, 6) if mapped else None,
            })
    return samples, dict(truth), observations


def extract_workbook(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    missing_sheets = [sheet_name for sheet_name in FULL_RESULT_SHEETS if sheet_name not in wb.sheetnames]
    if missing_sheets:
        raise ValueError(f"Workbook {workbook_path} missing required result sheets: {', '.join(missing_sheets)}")
    all_samples: dict[str, dict[str, Any]] = {}
    truth: dict[str, dict[str, list[str]]] = {}
    observations_by_sample: dict[str, list[dict[str, Any]]] = {}
    superseded_samples: list[dict[str, str]] = []
    for sheet_name in FULL_RESULT_SHEETS:
        sheet_samples, sheet_truth, sheet_observations = extract_sheet(wb[sheet_name])
        sheet_sample_ids = {sample["gs_id"] for sample in sheet_samples}
        for sample in sheet_samples:
            sample_id = sample["gs_id"]
            if sample_id in all_samples:
                superseded_samples.append({
                    "gs_id": sample_id,
                    "superseded_sheet": all_samples[sample_id]["sheet"],
                    "retained_sheet": sample["sheet"],
                    "reason": "later full-result sheet supersedes earlier occurrence for same gs_id",
                })
            all_samples[sample_id] = sample
            truth.pop(sample_id, None)
            observations_by_sample[sample_id] = []
        for sample_id, calls in sheet_truth.items():
            if sample_id in sheet_sample_ids:
                truth[sample_id] = calls
        for observation in sheet_observations:
            sample_id = observation["sample_id"]
            if sample_id in sheet_sample_ids:
                observations_by_sample.setdefault(sample_id, []).append(observation)
    samples = [all_samples[key] for key in sorted(all_samples)]
    observations = [observation for rows in observations_by_sample.values() for observation in rows]
    validate_extracted_dataset(workbook_path, samples, truth, observations)
    prompt_input = {
        "schema_version": 1,
        "dataset": workbook_path.name,
        "instructions": {
            "truth_blinded": True,
            "report_loci": REPORT_LOCI,
            "hidden_truth_source": "human-curated haplotype rows are excluded from prompt input",
            "duplicate_sample_policy": DUPLICATE_SAMPLE_POLICY,
        },
        "superseded_samples": superseded_samples,
        "samples": samples,
        "observations": sorted(observations, key=lambda row: (row["sample_id"], row["report_locus"], row["genotype"])),
    }
    return {"samples": samples, "truth_calls": truth, "prompt_input": prompt_input}


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def render_prompt_text(prompt_template: str, prompt_input: dict[str, Any]) -> str:
    placeholder_count = prompt_template.count(PROMPT_INPUT_PLACEHOLDER)
    if placeholder_count != 1:
        raise ValueError(f"Prompt template must contain exactly one {PROMPT_INPUT_PLACEHOLDER} placeholder")
    prompt_json = json.dumps(prompt_input, indent=2, sort_keys=True)
    return prompt_template.replace(PROMPT_INPUT_PLACEHOLDER, prompt_json)


def openai_request_payload(
    system_prompt: str,
    user_prompt: str,
    model: str,
    max_output_tokens: int,
    reasoning_effort: str | None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "model": model,
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_output_tokens": max_output_tokens,
        "temperature": 0,
    }
    if reasoning_effort:
        payload["reasoning"] = {"effort": reasoning_effort}
    return payload


def extract_response_text(response: dict[str, Any]) -> str:
    if "output_text" in response and isinstance(response["output_text"], str):
        return response["output_text"]
    chunks = []
    for item in response.get("output", []):
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []):
            if not isinstance(content, dict):
                continue
            if content.get("type") in {"output_text", "text"} and isinstance(content.get("text"), str):
                chunks.append(content["text"])
    if chunks:
        return "\n".join(chunks)
    raise ValueError("OpenAI response did not contain output text")


def redact_sensitive_text(text: Any, secrets: list[str | None] | None = None) -> str:
    redacted = "" if text is None else str(text)
    redacted = re.sub(
        r"(?i)(authorization\s*:\s*bearer\s+)[^\s,;\"']+",
        rf"\1{SECRET_REDACTION_MARKER}",
        redacted,
    )
    redacted = re.sub(r"(?i)(bearer\s+)[^\s,;\"']+", rf"\1{SECRET_REDACTION_MARKER}", redacted)
    redacted = re.sub(
        r"(?i)((?:api[_-]?key|token)\s*[:=]\s*)[^\s,;\"']+",
        rf"\1{SECRET_REDACTION_MARKER}",
        redacted,
    )
    for secret in secrets or []:
        if secret:
            redacted = redacted.replace(secret, SECRET_REDACTION_MARKER)
    return redacted


def call_openai_responses(api_key: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        error_body = redact_sensitive_text(exc.read().decode("utf-8", errors="replace"), [api_key])
        raise RuntimeError(f"OpenAI Responses API request failed with HTTP {exc.code}: {error_body}") from exc


def string_field(record: dict[str, Any], key: str, prefix: str, errors: list[str]) -> str | None:
    if key not in record:
        return None
    value = record[key]
    if not isinstance(value, str):
        errors.append(f"{prefix}.{key} must be a string")
        return None
    return value


def string_list_field(record: dict[str, Any], key: str, prefix: str, errors: list[str]) -> list[str]:
    if key not in record:
        return []
    value = record[key]
    if not isinstance(value, list):
        errors.append(f"{prefix}.{key} must be a list")
        return []
    strings = []
    for item in value:
        if not isinstance(item, str):
            errors.append(f"{prefix}.{key} must contain strings")
            continue
        strings.append(item)
    return strings


def validate_model_output(output: Any, prompt_input: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not isinstance(output, dict):
        return ["model output must be a JSON object"]
    if not isinstance(prompt_input, dict):
        prompt_input = {}

    required_top_level = {"schema_version", "prompt_version", "haplotype_definitions", "sample_calls", "unresolved"}
    for key in sorted(required_top_level - set(output)):
        errors.append(f"missing top-level key: {key}")
    for key in sorted(set(output) - required_top_level):
        errors.append(f"extra top-level key: {key}")
    schema_version = output.get("schema_version")
    if not (type(schema_version) is int and schema_version == 1):
        errors.append("schema_version must be integer 1")
    if output.get("prompt_version") != PROMPT_VERSION:
        errors.append(f"prompt_version must be {PROMPT_VERSION}")

    samples = prompt_input.get("samples", [])
    sample_ids = {sample.get("gs_id") for sample in samples if isinstance(sample, dict) and sample.get("gs_id")}
    instructions = prompt_input.get("instructions", {})
    report_loci = instructions.get("report_loci", []) if isinstance(instructions, dict) else []
    known_loci = {locus for locus in report_loci if isinstance(locus, str)}
    observations = prompt_input.get("observations", [])
    known_genotypes = {
        observation.get("genotype")
        for observation in observations
        if isinstance(observation, dict) and observation.get("genotype")
    }
    known_loci.update(
        observation.get("report_locus")
        for observation in observations
        if isinstance(observation, dict) and isinstance(observation.get("report_locus"), str)
    )

    definitions = output.get("haplotype_definitions")
    labels_by_locus: dict[str, set[str]] = defaultdict(set)
    if not isinstance(definitions, list):
        errors.append("haplotype_definitions must be a list")
    else:
        definition_required = {"locus", "label", "supporting_genotypes", "seed_samples", "confidence", "rationale"}
        for index, definition in enumerate(definitions):
            prefix = f"haplotype_definitions[{index}]"
            if not isinstance(definition, dict):
                errors.append(f"{prefix} must be an object")
                continue
            for key in sorted(definition_required - set(definition)):
                errors.append(f"{prefix} missing required key: {key}")
            locus = string_field(definition, "locus", prefix, errors)
            label = string_field(definition, "label", prefix, errors)
            if isinstance(locus, str) and isinstance(label, str):
                labels_by_locus[locus].add(label)
            if locus is not None and locus not in known_loci:
                errors.append(f"{prefix}.locus unknown locus: {locus}")
            for genotype in string_list_field(definition, "supporting_genotypes", prefix, errors):
                if genotype not in known_genotypes:
                    errors.append(f"{prefix}.supporting_genotypes unknown genotype: {genotype}")
            for sample_id in string_list_field(definition, "seed_samples", prefix, errors):
                if sample_id not in sample_ids:
                    errors.append(f"{prefix}.seed_samples unknown sample: {sample_id}")
            confidence = string_field(definition, "confidence", prefix, errors)
            if confidence is not None and confidence not in {"high", "medium", "low"}:
                errors.append(f"{prefix}.confidence must be high, medium, or low")
            string_field(definition, "rationale", prefix, errors)

    calls = output.get("sample_calls")
    if not isinstance(calls, list):
        errors.append("sample_calls must be a list")
    else:
        call_required = {
            "sample_id",
            "locus",
            "h1",
            "h2",
            "status",
            "h1_supporting_genotypes",
            "h2_supporting_genotypes",
            "rationale",
        }
        for index, call in enumerate(calls):
            prefix = f"sample_calls[{index}]"
            if not isinstance(call, dict):
                errors.append(f"{prefix} must be an object")
                continue
            for key in sorted(call_required - set(call)):
                errors.append(f"{prefix} missing required key: {key}")
            sample_id = string_field(call, "sample_id", prefix, errors)
            locus = string_field(call, "locus", prefix, errors)
            if sample_id is not None and sample_id not in sample_ids:
                errors.append(f"{prefix}.sample_id unknown sample: {sample_id}")
            if locus is not None and locus not in known_loci:
                errors.append(f"{prefix}.locus unknown locus: {locus}")
            status = string_field(call, "status", prefix, errors)
            if status is not None and status not in {"called", "partial", "unresolved"}:
                errors.append(f"{prefix}.status must be called, partial, or unresolved")
            for haplotype_key in ("h1", "h2"):
                label = string_field(call, haplotype_key, prefix, errors)
                if label is not None and locus is not None and label != "?" and label not in labels_by_locus.get(locus, set()):
                    errors.append(f"{prefix}.{haplotype_key} unknown haplotype label for locus {locus}: {label}")
            for genotype_key in ("h1_supporting_genotypes", "h2_supporting_genotypes"):
                for genotype in string_list_field(call, genotype_key, prefix, errors):
                    if genotype not in known_genotypes:
                        errors.append(f"{prefix}.{genotype_key} unknown genotype: {genotype}")
            string_field(call, "rationale", prefix, errors)

    unresolved = output.get("unresolved")
    if not isinstance(unresolved, list):
        errors.append("unresolved must be a list")
    else:
        unresolved_required = {"sample_id", "locus", "reason", "evidence_summary"}
        for index, item in enumerate(unresolved):
            prefix = f"unresolved[{index}]"
            if not isinstance(item, dict):
                errors.append(f"{prefix} must be an object")
                continue
            for key in sorted(unresolved_required - set(item)):
                errors.append(f"{prefix} missing required key: {key}")
            sample_id = string_field(item, "sample_id", prefix, errors)
            locus = string_field(item, "locus", prefix, errors)
            if sample_id is not None and sample_id not in sample_ids:
                errors.append(f"{prefix}.sample_id unknown sample: {sample_id}")
            if locus is not None and locus not in known_loci:
                errors.append(f"{prefix}.locus unknown locus: {locus}")
            string_field(item, "reason", prefix, errors)
            string_field(item, "evidence_summary", prefix, errors)

    return errors


def is_resolved_label(label: Any) -> bool:
    text = clean(label)
    return bool(text) and text != "?"


def call_index(output: dict[str, Any]) -> dict[str, dict[str, dict[str, Any]]]:
    indexed: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    calls = output.get("sample_calls", []) if isinstance(output, dict) else []
    if not isinstance(calls, list):
        return {}
    for call in calls:
        if not isinstance(call, dict):
            continue
        sample_id = clean(call.get("sample_id"))
        locus = clean(call.get("locus"))
        if sample_id and locus:
            indexed[sample_id][locus] = call
    return {sample_id: dict(locus_calls) for sample_id, locus_calls in sorted(indexed.items())}


def label_counter(labels: list[Any] | tuple[Any, ...]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for label in labels:
        text = clean(label)
        if text and text != "?":
            counts[text] += 1
    return counts


def max_weight_label_mapping(weights: dict[str, dict[str, int | float]]) -> dict[str, str]:
    positive_weights: dict[str, dict[str, float]] = defaultdict(dict)
    for pred_label, human_weights in weights.items():
        pred = clean(pred_label)
        if not pred or not isinstance(human_weights, dict):
            continue
        for human_label, weight in human_weights.items():
            human = clean(human_label)
            if not human:
                continue
            try:
                numeric_weight = float(weight)
            except (TypeError, ValueError):
                continue
            if numeric_weight > 0:
                positive_weights[pred][human] = positive_weights[pred].get(human, 0.0) + numeric_weight

    pred_labels = sorted(label for label, row in positive_weights.items() if row)
    human_labels = sorted({human for row in positive_weights.values() for human in row})
    if not pred_labels or not human_labels:
        return {}

    source = 0
    pred_offset = 1
    human_offset = pred_offset + len(pred_labels)
    sink = human_offset + len(human_labels)
    graph: list[list[dict[str, Any]]] = [[] for _ in range(sink + 1)]
    pred_nodes = {label: pred_offset + index for index, label in enumerate(pred_labels)}
    human_nodes = {label: human_offset + index for index, label in enumerate(human_labels)}

    def add_edge(start: int, end: int, capacity: int, cost: float, pred: str | None = None, human: str | None = None) -> None:
        forward = {"to": end, "rev": len(graph[end]), "cap": capacity, "cost": cost, "pred": pred, "human": human}
        reverse = {"to": start, "rev": len(graph[start]), "cap": 0, "cost": -cost, "pred": None, "human": None}
        graph[start].append(forward)
        graph[end].append(reverse)

    for pred in pred_labels:
        add_edge(source, pred_nodes[pred], 1, 0.0)
    for pred in pred_labels:
        for human in sorted(positive_weights[pred]):
            add_edge(pred_nodes[pred], human_nodes[human], 1, -positive_weights[pred][human], pred=pred, human=human)
    for human in human_labels:
        add_edge(human_nodes[human], sink, 1, 0.0)

    while True:
        distance = [float("inf")] * len(graph)
        parent: list[tuple[int, int] | None] = [None] * len(graph)
        in_queue = [False] * len(graph)
        distance[source] = 0.0
        queue: deque[int] = deque([source])
        in_queue[source] = True
        while queue:
            node = queue.popleft()
            in_queue[node] = False
            for edge_index, edge in enumerate(graph[node]):
                if edge["cap"] <= 0:
                    continue
                next_node = edge["to"]
                candidate = distance[node] + edge["cost"]
                if candidate < distance[next_node]:
                    distance[next_node] = candidate
                    parent[next_node] = (node, edge_index)
                    if not in_queue[next_node]:
                        queue.append(next_node)
                        in_queue[next_node] = True

        if parent[sink] is None or distance[sink] >= 0:
            break

        node = sink
        while node != source:
            prev_node, edge_index = parent[node]
            edge = graph[prev_node][edge_index]
            edge["cap"] -= 1
            graph[node][edge["rev"]]["cap"] += 1
            node = prev_node

    mapping: dict[str, str] = {}
    for pred in pred_labels:
        pred_node = pred_nodes[pred]
        for edge in graph[pred_node]:
            if edge.get("human") is not None and edge["cap"] == 0:
                mapping[pred] = edge["human"]
                break
    return mapping


def mapping_weights(output: dict[str, Any], truth: dict[str, dict[str, list[str]]], locus: str) -> dict[str, dict[str, int]]:
    calls = call_index(output)
    sample_counters: list[tuple[Counter[str], Counter[str]]] = []
    cooccurrence_weights: dict[str, Counter[str]] = defaultdict(Counter)
    for sample_id in sorted(truth):
        truth_slots = truth.get(sample_id, {}).get(locus)
        if not isinstance(truth_slots, list):
            continue
        call = calls.get(sample_id, {}).get(locus)
        if not isinstance(call, dict):
            continue
        if clean(call.get("status")).lower() == "unresolved":
            continue
        predicted_counter = label_counter([call.get("h1"), call.get("h2")])
        human_counter = label_counter(truth_slots[:2])
        if not predicted_counter or not human_counter:
            continue
        sample_counters.append((predicted_counter, human_counter))
        for predicted in sorted(predicted_counter):
            for human in sorted(human_counter):
                cooccurrence_weights[predicted][human] += min(predicted_counter[predicted], human_counter[human])

    inferred_mapping = max_weight_label_mapping(
        {predicted: dict(cooccurrence_weights[predicted]) for predicted in sorted(cooccurrence_weights)}
    )
    weights: dict[str, Counter[str]] = defaultdict(Counter)
    for predicted_counter, human_counter in sample_counters:
        remaining_predicted = Counter(predicted_counter)
        remaining_human = Counter(human_counter)
        for predicted in sorted(remaining_predicted):
            human = inferred_mapping.get(predicted)
            if human is None:
                continue
            matched = min(remaining_predicted[predicted], remaining_human.get(human, 0))
            if matched <= 0:
                continue
            weights[predicted][human] += matched
            remaining_predicted[predicted] -= matched
            remaining_human[human] -= matched

        residual_humans = []
        for human in sorted(remaining_human):
            residual_humans.extend([human] * remaining_human[human])
        for predicted in sorted(remaining_predicted):
            while remaining_predicted[predicted] > 0 and residual_humans:
                weights[predicted][residual_humans.pop(0)] += 1
                remaining_predicted[predicted] -= 1
    return {predicted: dict(weights[predicted]) for predicted in sorted(weights)}


def metric_rate(numerator: int, denominator: int) -> float:
    return round(numerator / denominator, 6) if denominator else 0.0


def false_merge_count(weights: dict[str, dict[str, int]]) -> int:
    return sum(1 for human_weights in weights.values() if sum(1 for weight in human_weights.values() if weight > 0) > 1)


def false_split_count(weights: dict[str, dict[str, int]]) -> int:
    humans_to_predictions: dict[str, set[str]] = defaultdict(set)
    for predicted, human_weights in weights.items():
        for human, weight in human_weights.items():
            if weight > 0:
                humans_to_predictions[human].add(predicted)
    return sum(1 for predictions in humans_to_predictions.values() if len(predictions) > 1)


def score_predictions(output: dict[str, Any], truth: dict[str, dict[str, list[str]]], loci: list[str] | None = None) -> dict[str, Any]:
    calls = call_index(output)
    sample_calls = output.get("sample_calls", []) if isinstance(output, dict) else []
    if not isinstance(sample_calls, list):
        sample_calls = []
    output_loci = {
        clean(call.get("locus"))
        for call in sample_calls
        if isinstance(call, dict) and clean(call.get("locus"))
    }
    truth_loci = {locus for sample_calls in truth.values() for locus in sample_calls}
    if loci is None:
        ordered_loci = [locus for locus in REPORT_LOCI if locus in truth_loci or locus in output_loci]
        ordered_loci.extend(sorted((truth_loci | output_loci) - set(ordered_loci)))
    else:
        ordered_loci = sorted(loci)

    overall = {
        "slot_hits": 0,
        "slot_total": 0,
        "pair_hits": 0,
        "pair_total": 0,
        "unresolved_count": 0,
        "call_total": 0,
        "false_merge_count": 0,
        "false_split_count": 0,
    }
    locus_scores: dict[str, Any] = {}

    for locus in ordered_loci:
        weights = mapping_weights(output, truth, locus)
        label_mapping = max_weight_label_mapping(weights)
        locus_counts = {
            "slot_hits": 0,
            "slot_total": 0,
            "pair_hits": 0,
            "pair_total": 0,
            "unresolved_count": 0,
            "call_total": 0,
            "false_merge_count": false_merge_count(weights),
            "false_split_count": false_split_count(weights),
        }

        for sample_id in sorted(truth):
            truth_slots = truth.get(sample_id, {}).get(locus)
            if not isinstance(truth_slots, list):
                continue
            human_counter = label_counter(truth_slots[:2])
            if not human_counter:
                continue
            call = calls.get(sample_id, {}).get(locus)
            predicted_slots = [call.get("h1"), call.get("h2")] if isinstance(call, dict) else []
            status = clean(call.get("status")).lower() if isinstance(call, dict) else ""
            is_unresolved = (
                not isinstance(call, dict)
                or status == "unresolved"
                or len(predicted_slots) < 2
                or any(not is_resolved_label(slot) for slot in predicted_slots[:2])
            )
            if is_unresolved:
                locus_counts["unresolved_count"] += 1
            if not isinstance(call, dict) or status == "unresolved":
                mapped_slots = []
            else:
                mapped_slots = [
                    label_mapping[clean(slot)]
                    for slot in predicted_slots[:2]
                    if is_resolved_label(slot) and clean(slot) in label_mapping
                ]
            predicted_counter = label_counter(mapped_slots)
            slot_hits = sum(min(predicted_counter[label], human_counter[label]) for label in human_counter)
            slot_total = sum(human_counter.values())
            locus_counts["slot_hits"] += slot_hits
            locus_counts["slot_total"] += slot_total
            locus_counts["pair_hits"] += int(predicted_counter == human_counter)
            locus_counts["pair_total"] += 1
            locus_counts["call_total"] += 1

        locus_score = {
            **locus_counts,
            "slot_concordance": metric_rate(locus_counts["slot_hits"], locus_counts["slot_total"]),
            "pair_concordance": metric_rate(locus_counts["pair_hits"], locus_counts["pair_total"]),
            "unresolved_rate": metric_rate(locus_counts["unresolved_count"], locus_counts["call_total"]),
            "label_mapping": {predicted: label_mapping[predicted] for predicted in sorted(label_mapping)},
        }
        locus_scores[locus] = locus_score
        for key in overall:
            overall[key] += locus_counts[key]

    overall_score = {
        **overall,
        "slot_concordance": metric_rate(overall["slot_hits"], overall["slot_total"]),
        "pair_concordance": metric_rate(overall["pair_hits"], overall["pair_total"]),
        "unresolved_rate": metric_rate(overall["unresolved_count"], overall["call_total"]),
    }
    return {"schemaVersion": 1, "overall": overall_score, "loci": locus_scores}


def runtime_identity() -> dict[str, Any]:
    return {
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "executable": sys.executable,
        "condaPrefix": os.environ.get("CONDA_PREFIX"),
        "container": os.environ.get("container"),
    }


def resolved_extract_options(workbook: Path, output_dir: Path) -> dict[str, Any]:
    return {
        "workbook": str(workbook),
        "outputDir": str(output_dir),
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
        },
    }


def resolved_render_prompt_options(iteration_dir: Path, prompt: Path) -> dict[str, Any]:
    return {
        "iterationDir": str(iteration_dir),
        "prompt": str(prompt),
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
        },
    }


def resolved_import_output_options(iteration_dir: Path, model_output: Path) -> dict[str, Any]:
    return {
        "iterationDir": str(iteration_dir),
        "modelOutput": str(model_output),
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
        },
    }


def resolved_score_options(iteration_dir: Path) -> dict[str, Any]:
    return {
        "iterationDir": str(iteration_dir),
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
        },
    }


def resolved_run_openai_options(
    iteration_dir: Path,
    model: str,
    max_output_tokens: int,
    reasoning_effort: str | None,
) -> dict[str, Any]:
    return {
        "iterationDir": str(iteration_dir),
        "model": model,
        "maxOutputTokens": max_output_tokens,
        "reasoningEffort": reasoning_effort,
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
            "model": DEFAULT_OPENAI_MODEL,
            "maxOutputTokens": DEFAULT_OPENAI_MAX_OUTPUT_TOKENS,
            "reasoningEffort": DEFAULT_OPENAI_REASONING_EFFORT,
        },
    }


def resolved_run_iteration_options(
    workbook: Path,
    iteration_dir: Path,
    prompt: Path,
    model: str,
    run_provider: bool,
) -> dict[str, Any]:
    return {
        "workbook": str(workbook),
        "iterationDir": str(iteration_dir),
        "prompt": str(prompt),
        "model": model,
        "runProvider": run_provider,
        "defaults": {
            "workbook": str(DEFAULT_SNPRC_WORKBOOK),
            "prompt": str(DEFAULT_PROMPT),
            "outputRoot": str(DEFAULT_OUTPUT_ROOT),
            "reportLoci": REPORT_LOCI,
            "fullResultSheets": FULL_RESULT_SHEETS,
            "model": DEFAULT_OPENAI_MODEL,
            "maxOutputTokens": DEFAULT_OPENAI_MAX_OUTPUT_TOKENS,
            "reasoningEffort": DEFAULT_OPENAI_REASONING_EFFORT,
        },
    }


def iteration_output_paths(iteration_dir: Path, run_provider: bool) -> list[Path]:
    paths = [
        iteration_dir / "prompt_input.json",
        iteration_dir / "truth_calls.json",
        iteration_dir / "samples.json",
        iteration_dir / "rendered_prompt.md",
    ]
    if run_provider:
        paths.extend(
            [
                iteration_dir / "openai_response.json",
                iteration_dir / "model_output.json",
                iteration_dir / "model_output_validation.json",
                iteration_dir / "parsed_model_output.json",
                iteration_dir / "score.json",
                iteration_dir / "score.md",
            ]
        )
    return paths


def provider_artifact_paths(iteration_dir: Path) -> list[Path]:
    return [
        iteration_dir / "openai_response.json",
        iteration_dir / "model_output.json",
        iteration_dir / "model_output_validation.json",
        iteration_dir / "parsed_model_output.json",
        iteration_dir / "score.json",
        iteration_dir / "score.md",
    ]


def remove_existing_paths(paths: list[Path]) -> None:
    for path in paths:
        if path.exists():
            path.unlink()


def write_provenance(
    output_dir: Path,
    workflow: str,
    argv: list[str],
    inputs: list[Path],
    outputs: list[Path],
    started: float,
    options: dict[str, Any],
    status: str = "completed",
    stderr: str | None = None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "schemaVersion": 1,
        "workflowName": workflow,
        "toolName": TOOL_NAME,
        "toolVersion": TOOL_VERSION,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "argv": argv,
        "reproducibleShellCommand": " ".join(shlex.quote(part) for part in [sys.executable, *argv]),
        "options": options,
        "runtimeIdentity": runtime_identity(),
        "inputs": [file_record(path, "input") for path in inputs if path.exists()],
        "outputs": [file_record(path, "output") for path in outputs if path.exists()],
        "exitStatus": 0 if status == "completed" else 1,
        "wallTimeSeconds": round(time.time() - started, 3),
        "stderr": stderr,
        "status": status,
    }
    write_json(output_dir / f"{workflow}.provenance.json", payload)


def command_extract(args: argparse.Namespace) -> None:
    started = time.time()
    workbook = args.workbook.resolve()
    out = args.output_dir.resolve()
    extracted = extract_workbook(workbook)
    evidence_path = out / "prompt_input.json"
    truth_path = out / "truth_calls.json"
    samples_path = out / "samples.json"
    write_json(evidence_path, extracted["prompt_input"])
    write_json(truth_path, extracted["truth_calls"])
    write_json(samples_path, extracted["samples"])
    write_provenance(
        out,
        "extract",
        args.effective_argv,
        [workbook],
        [evidence_path, truth_path, samples_path],
        started,
        resolved_extract_options(workbook, out),
    )


def command_render_prompt(args: argparse.Namespace) -> None:
    started = time.time()
    iteration_dir = args.iteration_dir.resolve()
    prompt_path = args.prompt.resolve()
    prompt_input_path = iteration_dir / "prompt_input.json"
    output_path = iteration_dir / "rendered_prompt.md"
    prompt_template = prompt_path.read_text(encoding="utf-8")
    prompt_input = json.loads(prompt_input_path.read_text(encoding="utf-8"))
    rendered = render_prompt_text(prompt_template, prompt_input)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(rendered, encoding="utf-8")
    write_provenance(
        iteration_dir,
        "render-prompt",
        args.effective_argv,
        [prompt_path, prompt_input_path],
        [output_path],
        started,
        resolved_render_prompt_options(iteration_dir, prompt_path),
    )


def command_import_output(args: argparse.Namespace) -> None:
    started = time.time()
    iteration_dir = args.iteration_dir.resolve()
    model_output_path = args.model_output.resolve()
    prompt_input_path = iteration_dir / "prompt_input.json"
    validation_path = iteration_dir / "model_output_validation.json"
    parsed_output_path = iteration_dir / "parsed_model_output.json"
    options = resolved_import_output_options(iteration_dir, model_output_path)
    inputs = [prompt_input_path, model_output_path]

    try:
        prompt_input = json.loads(prompt_input_path.read_text(encoding="utf-8"))
        model_output = json.loads(model_output_path.read_text(encoding="utf-8"))
    except Exception as exc:
        message = f"model output read/parse failure: {exc}"
        if parsed_output_path.exists():
            parsed_output_path.unlink()
        write_json(
            validation_path,
            {
                "schemaVersion": 1,
                "accepted": False,
                "errors": [message],
            },
        )
        write_provenance(
            iteration_dir,
            "import-output",
            args.effective_argv,
            inputs,
            [validation_path],
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message) from exc

    try:
        errors = validate_model_output(model_output, prompt_input)
    except Exception as exc:
        errors = [f"model output validation raised unexpected error: {exc}"]
    validation = {
        "schemaVersion": 1,
        "accepted": not errors,
        "errors": errors,
    }
    write_json(validation_path, validation)
    if errors:
        if parsed_output_path.exists():
            parsed_output_path.unlink()
        message = "model output validation failed: " + "; ".join(errors)
        write_provenance(
            iteration_dir,
            "import-output",
            args.effective_argv,
            inputs,
            [validation_path],
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message)

    write_json(parsed_output_path, model_output)
    write_provenance(
        iteration_dir,
        "import-output",
        args.effective_argv,
        inputs,
        [validation_path, parsed_output_path],
        started,
        options,
    )


def markdown_score_report(score: dict[str, Any]) -> str:
    overall = score["overall"]
    lines = [
        "# Macaque MHC Prompt Lab Score",
        "",
        "## Overall",
        "",
        "| metric | value |",
        "| --- | ---: |",
        f"| slot_concordance | {overall['slot_concordance']:.3f} |",
        f"| pair_concordance | {overall['pair_concordance']:.3f} |",
        f"| unresolved_rate | {overall['unresolved_rate']:.3f} |",
        f"| slot_hits / slot_total | {overall['slot_hits']} / {overall['slot_total']} |",
        f"| pair_hits / pair_total | {overall['pair_hits']} / {overall['pair_total']} |",
        f"| false_merge_count | {overall['false_merge_count']} |",
        f"| false_split_count | {overall['false_split_count']} |",
        "",
        "## Per Locus",
        "",
        "| locus | slot_concordance | pair_concordance | unresolved_rate | slots | pairs | false_merges | false_splits |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for locus in sorted(score["loci"]):
        row = score["loci"][locus]
        lines.append(
            "| "
            + " | ".join(
                [
                    locus,
                    f"{row['slot_concordance']:.3f}",
                    f"{row['pair_concordance']:.3f}",
                    f"{row['unresolved_rate']:.3f}",
                    f"{row['slot_hits']} / {row['slot_total']}",
                    f"{row['pair_hits']} / {row['pair_total']}",
                    str(row["false_merge_count"]),
                    str(row["false_split_count"]),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def command_score(args: argparse.Namespace) -> None:
    started = time.time()
    iteration_dir = args.iteration_dir.resolve()
    truth_path = iteration_dir / "truth_calls.json"
    parsed_output_path = iteration_dir / "parsed_model_output.json"
    score_path = iteration_dir / "score.json"
    report_path = iteration_dir / "score.md"
    options = resolved_score_options(iteration_dir)
    inputs = [truth_path, parsed_output_path]
    try:
        truth = json.loads(truth_path.read_text(encoding="utf-8"))
        output = json.loads(parsed_output_path.read_text(encoding="utf-8"))
        score = score_predictions(output, truth)
        write_json(score_path, score)
        report_path.write_text(markdown_score_report(score), encoding="utf-8")
    except Exception as exc:
        message = f"score failed: {exc}"
        for stale_path in (score_path, report_path):
            if stale_path.exists():
                stale_path.unlink()
        write_provenance(
            iteration_dir,
            "score",
            args.effective_argv,
            inputs,
            [],
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message) from exc
    write_provenance(
        iteration_dir,
        "score",
        args.effective_argv,
        inputs,
        [score_path, report_path],
        started,
        options,
    )


def command_run_openai(args: argparse.Namespace) -> None:
    started = time.time()
    iteration_dir = args.iteration_dir.resolve()
    rendered_prompt_path = iteration_dir / "rendered_prompt.md"
    openai_response_path = iteration_dir / "openai_response.json"
    model_output_path = iteration_dir / "model_output.json"
    options = resolved_run_openai_options(
        iteration_dir,
        args.model,
        args.max_output_tokens,
        args.reasoning_effort,
    )
    inputs = [rendered_prompt_path]

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        message = "OPENAI_API_KEY is required for run-openai"
        remove_existing_paths([openai_response_path, model_output_path])
        write_provenance(
            iteration_dir,
            "run-openai",
            args.effective_argv,
            inputs,
            [],
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message)

    remove_existing_paths([openai_response_path, model_output_path])

    try:
        rendered_prompt = rendered_prompt_path.read_text(encoding="utf-8")
        payload = openai_request_payload(
            OPENAI_SYSTEM_PROMPT,
            rendered_prompt,
            model=args.model,
            max_output_tokens=args.max_output_tokens,
            reasoning_effort=args.reasoning_effort,
        )
        response = call_openai_responses(api_key, payload)
        write_json(openai_response_path, response)
        response_text = extract_response_text(response)
        model_output = json.loads(response_text)
        write_json(model_output_path, model_output)
    except Exception as exc:
        message = redact_sensitive_text(f"run-openai failed: {exc}", [api_key])
        remove_existing_paths([model_output_path])
        write_provenance(
            iteration_dir,
            "run-openai",
            args.effective_argv,
            inputs,
            [openai_response_path],
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message) from exc

    write_provenance(
        iteration_dir,
        "run-openai",
        args.effective_argv,
        inputs,
        [openai_response_path, model_output_path],
        started,
        options,
    )


def run_cli(argv: list[str]) -> int:
    return main(argv)


def run_iteration(
    workbook: Path,
    iteration_dir: Path,
    prompt_path: Path | None,
    run_provider: bool,
    model: str,
) -> int:
    prompt = prompt_path if prompt_path is not None else DEFAULT_PROMPT
    run_cli(["extract", "--workbook", str(workbook), "--output-dir", str(iteration_dir)])
    run_cli(["render-prompt", "--iteration-dir", str(iteration_dir), "--prompt", str(prompt)])
    if run_provider:
        run_cli(["run-openai", "--iteration-dir", str(iteration_dir), "--model", model])
        run_cli(
            [
                "import-output",
                "--iteration-dir",
                str(iteration_dir),
                "--model-output",
                str(iteration_dir / "model_output.json"),
            ]
        )
        run_cli(["score", "--iteration-dir", str(iteration_dir)])
    return 0


def command_run_iteration(args: argparse.Namespace) -> None:
    started = time.time()
    workbook = args.workbook.resolve()
    iteration_dir = args.iteration_dir.resolve()
    prompt = args.prompt.resolve()
    options = resolved_run_iteration_options(workbook, iteration_dir, prompt, args.model, args.run_provider)
    inputs = [workbook, prompt]
    outputs = iteration_output_paths(iteration_dir, args.run_provider)
    if args.run_provider:
        remove_existing_paths(provider_artifact_paths(iteration_dir))

    try:
        run_iteration(
            workbook=workbook,
            iteration_dir=iteration_dir,
            prompt_path=prompt,
            run_provider=args.run_provider,
            model=args.model,
        )
    except SystemExit as exc:
        message = f"run-iteration failed: {exc}"
        write_provenance(
            iteration_dir,
            "run-iteration",
            args.effective_argv,
            inputs,
            outputs,
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise
    except Exception as exc:
        message = f"run-iteration failed: {exc}"
        write_provenance(
            iteration_dir,
            "run-iteration",
            args.effective_argv,
            inputs,
            outputs,
            started,
            options,
            status="failed",
            stderr=message,
        )
        raise SystemExit(message) from exc

    write_provenance(
        iteration_dir,
        "run-iteration",
        args.effective_argv,
        inputs,
        outputs,
        started,
        options,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    extract = sub.add_parser("extract", help="Extract blinded prompt input and held-out truth from an SNPRC workbook.")
    extract.add_argument("--workbook", type=Path, default=DEFAULT_SNPRC_WORKBOOK)
    extract.add_argument("--output-dir", type=Path, required=True)
    extract.set_defaults(func=command_extract)
    render = sub.add_parser("render-prompt", help="Render the generalist prompt with blinded prompt input.")
    render.add_argument("--iteration-dir", type=Path, required=True)
    render.add_argument("--prompt", type=Path, default=DEFAULT_PROMPT)
    render.set_defaults(func=command_render_prompt)
    import_output = sub.add_parser("import-output", help="Validate and import a model output JSON file.")
    import_output.add_argument("--iteration-dir", type=Path, required=True)
    import_output.add_argument("--model-output", type=Path, required=True)
    import_output.set_defaults(func=command_import_output)
    score = sub.add_parser("score", help="Score parsed model output against held-out human calls.")
    score.add_argument("--iteration-dir", type=Path, required=True)
    score.set_defaults(func=command_score)
    run_openai = sub.add_parser("run-openai", help="Call OpenAI Responses API with rendered_prompt.md.")
    run_openai.add_argument("--iteration-dir", type=Path, required=True)
    run_openai.add_argument("--model", default=DEFAULT_OPENAI_MODEL)
    run_openai.add_argument("--max-output-tokens", type=int, default=DEFAULT_OPENAI_MAX_OUTPUT_TOKENS)
    run_openai.add_argument("--reasoning-effort", default=DEFAULT_OPENAI_REASONING_EFFORT)
    run_openai.set_defaults(func=command_run_openai)
    run_iteration_parser = sub.add_parser(
        "run-iteration",
        help="Extract, render, optionally run provider, import, and score one iteration.",
    )
    run_iteration_parser.add_argument("--workbook", type=Path, default=DEFAULT_SNPRC_WORKBOOK)
    run_iteration_parser.add_argument("--iteration-dir", type=Path, required=True)
    run_iteration_parser.add_argument("--prompt", type=Path, default=DEFAULT_PROMPT)
    run_iteration_parser.add_argument("--model", default=DEFAULT_OPENAI_MODEL)
    run_iteration_parser.add_argument("--run-provider", action="store_true")
    run_iteration_parser.set_defaults(func=command_run_iteration)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    args.effective_argv = sys.argv if argv is None else [str(Path(__file__)), *argv]
    args.func(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
