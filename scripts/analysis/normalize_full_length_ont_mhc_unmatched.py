#!/usr/bin/env python3
"""Trim full-length ONT MHC unmatched clusters from mapping boundaries.

Prototype workflow:
1. Read unmatched workbook rows from an existing .lungfishgenotype bundle.
2. Re-run the captured minimap2 mapping shape for selected sample cluster FASTAs.
3. Trim each unmatched cluster sequence to the best minimap2 hit's target interval.
4. Optionally run blastn on the trimmed sequences for post-trim closest-reference metadata.
5. Write detail/pivot reports and provenance for the derived output directory.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import platform
import re
import shlex
import shutil
import subprocess
import sys
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


TOOL_NAME = "normalize_full_length_ont_mhc_unmatched.py"
TOOL_VERSION = "2026-07-08.3"
DEFAULT_WORKBOOK_SHEET = "MHC-like Unmatched Clusters"
DEFAULT_BLAST_EVALUE = "1e-20"


@dataclass(frozen=True)
class FastaRecord:
    name: str
    sequence: str


@dataclass(frozen=True)
class WorkbookRow:
    unmatched_sequence_id: str
    sample: str
    cluster: str
    cluster_reads: int
    closest_match_id: str
    match_class: str
    sequence: str
    source_values: dict[str, str]


@dataclass(frozen=True)
class MinimapHit:
    cluster: str
    allele: str
    cluster_start: int
    cluster_end: int
    snps: int
    matched_bases: int
    indel_bases: int
    is_reverse: bool = False

    @property
    def score(self) -> int:
        return self.matched_bases - 10 * self.indel_bases - 100 * self.snps


@dataclass(frozen=True)
class BlastHit:
    query_id: str
    closest_reference: str
    percent_identity: float
    aligned_bases: int
    mismatches: int
    gap_opens: int
    qstart: int
    qend: int
    sstart: int
    send: int
    evalue: float
    bitscore: float
    query_length: int
    subject_length: int

    @property
    def query_coverage(self) -> float:
        if self.query_length <= 0:
            return 0.0
        return self.aligned_bases / self.query_length * 100.0


@dataclass(frozen=True)
class NormalizedRow:
    original: WorkbookRow
    raw_sequence_id: str
    trimmed_sequence_id: str
    raw_length: int
    trimmed_length: int
    trim_start: int | None
    trim_end: int | None
    trim_source: str
    minimap_allele: str
    mapping_closest_match_id: str
    mapping_match_class: str
    mapping_nucleotides_different: int | None
    minimap_snps: int | None
    minimap_indel_bases: int | None
    minimap_matched_bases: int | None
    minimap_score: int | None
    trimmed_sequence: str
    blast_hit: BlastHit | None


def cigar_target_span(cigar: str) -> int:
    span = 0
    for count_text, op in re.findall(r"(\d+)([MIDNSHP=X])", cigar):
        if op in {"M", "D", "N", "=", "X"}:
            span += int(count_text)
    return span


def parse_cigar_counts(cigar: str) -> tuple[int, int, int]:
    snps = 0
    matched = 0
    indels = 0
    for count_text, op in re.findall(r"(\d+)([MIDNSHP=X])", cigar):
        count = int(count_text)
        if op == "X":
            snps += count
        elif op in {"=", "M"}:
            matched += count
        elif op in {"I", "D"}:
            indels += count
    return snps, matched, indels


def parse_minimap_sam_hit(line: str) -> MinimapHit | None:
    if not line or line.startswith("@"):
        return None
    fields = line.rstrip("\n").split("\t")
    if len(fields) < 6:
        return None
    try:
        flag = int(fields[1])
        pos = int(fields[3])
    except ValueError:
        return None
    cluster = fields[2]
    cigar = fields[5]
    if flag & 4 or cluster == "*" or cigar == "*" or pos <= 0:
        return None
    span = cigar_target_span(cigar)
    if span <= 0:
        return None
    allele = fields[0].split()[0]
    snps, matched, indels = parse_cigar_counts(cigar)
    return MinimapHit(
        cluster=cluster,
        allele=allele,
        cluster_start=pos,
        cluster_end=pos + span - 1,
        snps=snps,
        matched_bases=matched,
        indel_bases=indels,
        is_reverse=bool(flag & 16),
    )


def best_minimap_hit(hits: Iterable[MinimapHit]) -> MinimapHit | None:
    values = list(hits)
    if not values:
        return None
    return sorted(
        values,
        key=lambda hit: (
            hit.snps,
            hit.indel_bases,
            -hit.matched_bases,
            -hit.score,
            hit.allele,
        ),
    )[0]


def mapping_match_class(hit: MinimapHit | None) -> str:
    if hit is None:
        return ""
    return "extension" if hit.snps == 0 else "snp-different"


def mapping_nucleotides_different(hit: MinimapHit | None) -> int | None:
    if hit is None:
        return None
    return 0 if hit.snps == 0 else hit.snps


def mapping_closest_match_id(hit: MinimapHit | None) -> str:
    if hit is None:
        return ""
    if hit.snps == 0:
        return f"{hit.allele}_extension"
    return f"{hit.allele}_{hit.snps}SNP"


def unmatched_sequence_id(sequence: str) -> str:
    normalized = sequence.strip().upper().encode("utf-8")
    bytes16 = bytearray(hashlib.sha256(normalized).digest()[:16])
    bytes16[6] = (bytes16[6] & 0x0F) | 0x50
    bytes16[8] = (bytes16[8] & 0x3F) | 0x80
    return str(uuid.UUID(bytes=bytes(bytes16)))


def reverse_complement(sequence: str) -> str:
    translation = str.maketrans("ACGTNacgtn", "TGCANtgcan")
    return sequence.translate(translation)[::-1].upper()


def read_fasta(path: Path) -> dict[str, FastaRecord]:
    records: dict[str, FastaRecord] = {}
    name: str | None = None
    parts: list[str] = []
    with path.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name is not None:
                    records[name] = FastaRecord(name=name, sequence="".join(parts).upper())
                name = line[1:].split()[0]
                parts = []
            else:
                parts.append(line)
    if name is not None:
        records[name] = FastaRecord(name=name, sequence="".join(parts).upper())
    return records


def write_fasta(records: Iterable[tuple[str, str]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for name, sequence in records:
            handle.write(f">{name}\n")
            for index in range(0, len(sequence), 80):
                handle.write(sequence[index : index + 80] + "\n")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_record(path: Path, role: str, file_format: str | None = None) -> dict[str, object]:
    record: dict[str, object] = {
        "path": str(path.resolve()),
        "role": role,
        "sizeBytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }
    if file_format:
        record["format"] = file_format
    return record


def load_manifest(bundle: Path) -> dict[str, object]:
    manifest = bundle / "genotype-result.json"
    if not manifest.is_file():
        raise FileNotFoundError(f"missing genotype-result.json in {bundle}")
    return json.loads(manifest.read_text(encoding="utf-8"))


def default_workbook_path(bundle: Path) -> Path:
    manifest = load_manifest(bundle)
    primary = manifest.get("primaryWorkbookPath")
    if not isinstance(primary, str) or not primary:
        raise ValueError("genotype-result.json lacks primaryWorkbookPath")
    return bundle / primary


def mapping_reference_from_provenance(bundle: Path) -> Path:
    manifest = load_manifest(bundle)
    provenance_path = manifest.get("provenancePath")
    if not isinstance(provenance_path, str) or not provenance_path:
        raise ValueError("genotype-result.json lacks provenancePath")
    provenance = json.loads((bundle / provenance_path).read_text(encoding="utf-8"))
    for step in provenance.get("steps", []):
        if not isinstance(step, dict):
            continue
        tool_name = str(step.get("toolName", ""))
        argv = step.get("argv") or step.get("command")
        if "minimap2" not in tool_name and not (isinstance(argv, list) and argv and "minimap2" in str(argv[0])):
            continue
        if isinstance(argv, list) and len(argv) >= 2:
            candidate = Path(str(argv[-1]))
            if candidate.exists():
                return candidate
    raise ValueError("could not resolve minimap2 mapping reference from provenance")


def load_workbook_rows(workbook_path: Path, sheet_name: str, samples: set[str] | None) -> list[WorkbookRow]:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError("openpyxl is required to read the genotype workbook") from error

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"workbook lacks sheet {sheet_name!r}; available: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    header = [str(value or "") for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = {"unmatched_sequence_id", "sample", "cluster", "cluster_reads", "sequence"}
    missing = required.difference(header)
    if missing:
        raise ValueError(f"sheet {sheet_name!r} lacks required columns: {sorted(missing)}")

    rows: list[WorkbookRow] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        source = {key: "" if value is None else str(value) for key, value in zip(header, values)}
        sample = source.get("sample", "")
        if samples is not None and sample not in samples:
            continue
        sequence = source.get("sequence", "").strip().upper()
        if not sequence:
            continue
        try:
            reads = int(float(source.get("cluster_reads", "0") or "0"))
        except ValueError:
            reads = 0
        rows.append(
            WorkbookRow(
                unmatched_sequence_id=source.get("unmatched_sequence_id", ""),
                sample=sample,
                cluster=source.get("cluster", ""),
                cluster_reads=reads,
                closest_match_id=source.get("closest_match_id", ""),
                match_class=source.get("match_class", ""),
                sequence=sequence,
                source_values=source,
            )
        )
    return rows


def run_command(argv: list[str], cwd: Path, stdout_path: Path) -> dict[str, object]:
    started = dt.datetime.now(dt.UTC)
    start_time = time.monotonic()
    with stdout_path.open("w", encoding="utf-8") as stdout_handle:
        result = subprocess.run(
            argv,
            cwd=str(cwd),
            text=True,
            stdout=stdout_handle,
            stderr=subprocess.PIPE,
            check=False,
        )
    completed = dt.datetime.now(dt.UTC)
    return {
        "toolName": Path(argv[0]).name,
        "toolVersion": tool_version(argv[0]),
        "argv": argv,
        "command": argv,
        "reproducibleCommand": shlex.join(argv),
        "startedAt": started.isoformat().replace("+00:00", "Z"),
        "completedAt": completed.isoformat().replace("+00:00", "Z"),
        "exitStatus": result.returncode,
        "exitCode": result.returncode,
        "wallTimeSeconds": time.monotonic() - start_time,
        "stderr": result.stderr,
        "outputs": [str(stdout_path.resolve())],
    }


def tool_version(executable: str) -> str:
    try:
        if Path(executable).name in {"blastn", "makeblastdb"}:
            result = subprocess.run([executable, "-version"], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
            return result.stdout.splitlines()[0].strip() if result.stdout else "unknown"
        result = subprocess.run([executable, "--version"], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
        return (result.stdout or result.stderr).splitlines()[0].strip()
    except Exception:
        return "unknown"


def parse_sam_hits(sam_text: str) -> dict[str, MinimapHit]:
    grouped: dict[str, list[MinimapHit]] = defaultdict(list)
    for line in sam_text.splitlines():
        hit = parse_minimap_sam_hit(line)
        if hit is not None:
            grouped[hit.cluster].append(hit)
    return {cluster: best for cluster, hits in grouped.items() if (best := best_minimap_hit(hits)) is not None}


def run_minimap_for_samples(
    bundle: Path,
    samples: list[str],
    mapping_reference: Path,
    output_dir: Path,
    threads: int,
    minimap2: str,
) -> tuple[dict[tuple[str, str], MinimapHit], list[dict[str, object]], list[Path]]:
    steps: list[dict[str, object]] = []
    input_paths: list[Path] = [mapping_reference]
    hits: dict[tuple[str, str], MinimapHit] = {}
    minimap_dir = output_dir / "intermediates" / "minimap2"
    minimap_dir.mkdir(parents=True, exist_ok=True)
    for sample in samples:
        cluster_fasta = bundle / "samples" / sample / "savont" / f"{sample}.savont-clusters.fasta"
        if not cluster_fasta.is_file():
            raise FileNotFoundError(f"missing sample cluster FASTA: {cluster_fasta}")
        input_paths.append(cluster_fasta)
        sam_path = minimap_dir / f"{sample}.clusters-vs-reference.sam"
        argv = [
            minimap2,
            "-a",
            "-x",
            "splice",
            "--eqx",
            "-t",
            str(max(1, threads)),
            "-N",
            "100",
            "--secondary=yes",
            str(cluster_fasta.resolve()),
            str(mapping_reference.resolve()),
        ]
        step = run_command(argv, minimap_dir, sam_path)
        step["inputs"] = [str(cluster_fasta.resolve()), str(mapping_reference.resolve())]
        if step["exitStatus"] != 0:
            raise RuntimeError(f"minimap2 failed for {sample}: {step['stderr']}")
        steps.append(step)
        for cluster, hit in parse_sam_hits(sam_path.read_text(encoding="utf-8")).items():
            hits[(sample, cluster)] = hit
    return hits, steps, input_paths


def trim_rows(rows: list[WorkbookRow], hits: dict[tuple[str, str], MinimapHit]) -> list[NormalizedRow]:
    normalized: list[NormalizedRow] = []
    for row in rows:
        hit = hits.get((row.sample, row.cluster))
        raw = row.sequence.upper()
        if hit is None:
            trimmed = raw
            normalized.append(
                NormalizedRow(
                    original=row,
                    raw_sequence_id=unmatched_sequence_id(raw),
                    trimmed_sequence_id=unmatched_sequence_id(trimmed),
                    raw_length=len(raw),
                    trimmed_length=len(trimmed),
                    trim_start=None,
                    trim_end=None,
                    trim_source="none-no-minimap-hit",
                    minimap_allele="",
                    mapping_closest_match_id="",
                    mapping_match_class="",
                    mapping_nucleotides_different=None,
                    minimap_snps=None,
                    minimap_indel_bases=None,
                    minimap_matched_bases=None,
                    minimap_score=None,
                    trimmed_sequence=trimmed,
                    blast_hit=None,
                )
            )
            continue
        start = max(1, min(hit.cluster_start, hit.cluster_end))
        end = min(len(raw), max(hit.cluster_start, hit.cluster_end))
        trimmed = raw[start - 1 : end] if start <= end else raw
        trim_source = "minimap2-target-interval"
        if hit.is_reverse:
            trimmed = reverse_complement(trimmed)
            trim_source = "minimap2-target-interval-reverse-complement"
        normalized.append(
            NormalizedRow(
                original=row,
                raw_sequence_id=unmatched_sequence_id(raw),
                trimmed_sequence_id=unmatched_sequence_id(trimmed),
                raw_length=len(raw),
                trimmed_length=len(trimmed),
                trim_start=start,
                trim_end=end,
                trim_source=trim_source,
                minimap_allele=hit.allele,
                mapping_closest_match_id=mapping_closest_match_id(hit),
                mapping_match_class=mapping_match_class(hit),
                mapping_nucleotides_different=mapping_nucleotides_different(hit),
                minimap_snps=hit.snps,
                minimap_indel_bases=hit.indel_bases,
                minimap_matched_bases=hit.matched_bases,
                minimap_score=hit.score,
                trimmed_sequence=trimmed,
                blast_hit=None,
            )
        )
    return normalized


def parse_blast_hits(tsv_text: str) -> dict[str, BlastHit]:
    grouped: dict[str, list[BlastHit]] = defaultdict(list)
    for line in tsv_text.splitlines():
        fields = line.split("\t")
        if len(fields) < 14:
            continue
        try:
            hit = BlastHit(
                query_id=fields[0],
                closest_reference=fields[1],
                percent_identity=float(fields[2]),
                aligned_bases=int(fields[3]),
                mismatches=int(fields[4]),
                gap_opens=int(fields[5]),
                qstart=int(fields[6]),
                qend=int(fields[7]),
                sstart=int(fields[8]),
                send=int(fields[9]),
                evalue=float(fields[10]),
                bitscore=float(fields[11]),
                query_length=int(fields[12]),
                subject_length=int(fields[13]),
            )
        except ValueError:
            continue
        grouped[hit.query_id].append(hit)
    return {
        query: sorted(
            hits,
            key=lambda hit: (
                hit.evalue,
                -hit.bitscore,
                -hit.query_coverage,
                -hit.percent_identity,
                -hit.aligned_bases,
                hit.closest_reference,
            ),
        )[0]
        for query, hits in grouped.items()
    }


def run_blast(
    normalized: list[NormalizedRow],
    blast_reference: Path,
    output_dir: Path,
    blastn: str,
    makeblastdb: str,
    evalue: str,
    threads: int,
) -> tuple[list[NormalizedRow], list[dict[str, object]], Path, Path]:
    blast_dir = output_dir / "intermediates" / "blastn"
    blast_dir.mkdir(parents=True, exist_ok=True)
    steps: list[dict[str, object]] = []
    db_dir = blast_dir / "reference-db"
    db_dir.mkdir(parents=True, exist_ok=True)
    db_reference = db_dir / "reference.fa"
    shutil.copy2(blast_reference, db_reference)
    db_prefix = db_dir / "mhc-reference"
    makeblastdb_log = db_dir / "makeblastdb.stdout.log"
    makeblastdb_argv = [
        makeblastdb,
        "-in",
        str(db_reference.resolve()),
        "-dbtype",
        "nucl",
        "-out",
        str(db_prefix.resolve()),
    ]
    makeblastdb_step = run_command(makeblastdb_argv, db_dir, makeblastdb_log)
    makeblastdb_step["inputs"] = [str(blast_reference.resolve()), str(db_reference.resolve())]
    makeblastdb_step["outputs"] = [str(path.resolve()) for path in sorted(db_dir.glob("mhc-reference.*"))] + [str(makeblastdb_log.resolve())]
    if makeblastdb_step["exitStatus"] != 0:
        raise RuntimeError(f"makeblastdb failed: {makeblastdb_step['stderr']}")
    steps.append(makeblastdb_step)

    query_fasta = blast_dir / "trimmed-unmatched-query.fasta"
    query_records = [
        (
            f"{row.original.sample}|{row.original.cluster}|{row.trimmed_sequence_id}",
            row.trimmed_sequence,
        )
        for row in normalized
    ]
    write_fasta(query_records, query_fasta)
    tsv_path = blast_dir / "trimmed-unmatched-blast.tsv"
    outfmt = "6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore qlen slen"
    argv = [
        blastn,
        "-query",
        str(query_fasta.resolve()),
        "-db",
        str(db_prefix.resolve()),
        "-task",
        "blastn",
        "-dust",
        "no",
        "-evalue",
        evalue,
        "-num_threads",
        str(max(1, threads)),
        "-outfmt",
        outfmt,
    ]
    step = run_command(argv, blast_dir, tsv_path)
    step["inputs"] = [str(query_fasta.resolve()), str(blast_reference.resolve())]
    if step["exitStatus"] != 0:
        raise RuntimeError(f"blastn failed: {step['stderr']}")
    steps.append(step)
    hits = parse_blast_hits(tsv_path.read_text(encoding="utf-8"))
    updated = []
    for row, (query_id, _) in zip(normalized, query_records):
        updated.append(
            NormalizedRow(
                original=row.original,
                raw_sequence_id=row.raw_sequence_id,
                trimmed_sequence_id=row.trimmed_sequence_id,
                raw_length=row.raw_length,
                trimmed_length=row.trimmed_length,
                trim_start=row.trim_start,
                trim_end=row.trim_end,
                trim_source=row.trim_source,
                minimap_allele=row.minimap_allele,
                mapping_closest_match_id=row.mapping_closest_match_id,
                mapping_match_class=row.mapping_match_class,
                mapping_nucleotides_different=row.mapping_nucleotides_different,
                minimap_snps=row.minimap_snps,
                minimap_indel_bases=row.minimap_indel_bases,
                minimap_matched_bases=row.minimap_matched_bases,
                minimap_score=row.minimap_score,
                trimmed_sequence=row.trimmed_sequence,
                blast_hit=hits.get(query_id),
            )
        )
    return updated, steps, query_fasta, tsv_path


def write_detail_csv(rows: list[NormalizedRow], path: Path) -> None:
    fields = [
        "trimmed_sequence_id",
        "raw_sequence_id",
        "sample",
        "cluster",
        "cluster_reads",
        "raw_length",
        "trimmed_length",
        "trim_start",
        "trim_end",
        "trim_source",
        "original_closest_match_id",
        "original_match_class",
        "minimap_allele",
        "mapping_closest_match_id",
        "mapping_match_class",
        "mapping_nucleotides_different",
        "minimap_snps",
        "minimap_indel_bases",
        "minimap_matched_bases",
        "minimap_score",
        "post_trim_blast_reference",
        "post_trim_percent_identity",
        "post_trim_query_coverage",
        "post_trim_aligned_bases",
        "post_trim_evalue",
        "post_trim_bitscore",
        "trimmed_sequence",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in sorted(rows, key=lambda item: (item.original.sample, item.original.cluster)):
            blast = row.blast_hit
            writer.writerow(
                {
                    "trimmed_sequence_id": row.trimmed_sequence_id,
                    "raw_sequence_id": row.raw_sequence_id,
                    "sample": row.original.sample,
                    "cluster": row.original.cluster,
                    "cluster_reads": row.original.cluster_reads,
                    "raw_length": row.raw_length,
                    "trimmed_length": row.trimmed_length,
                    "trim_start": row.trim_start or "",
                    "trim_end": row.trim_end or "",
                    "trim_source": row.trim_source,
                    "original_closest_match_id": row.original.closest_match_id,
                    "original_match_class": row.original.match_class,
                    "minimap_allele": row.minimap_allele,
                    "mapping_closest_match_id": row.mapping_closest_match_id,
                    "mapping_match_class": row.mapping_match_class,
                    "mapping_nucleotides_different": "" if row.mapping_nucleotides_different is None else row.mapping_nucleotides_different,
                    "minimap_snps": "" if row.minimap_snps is None else row.minimap_snps,
                    "minimap_indel_bases": "" if row.minimap_indel_bases is None else row.minimap_indel_bases,
                    "minimap_matched_bases": "" if row.minimap_matched_bases is None else row.minimap_matched_bases,
                    "minimap_score": "" if row.minimap_score is None else row.minimap_score,
                    "post_trim_blast_reference": blast.closest_reference if blast else "",
                    "post_trim_percent_identity": format_float(blast.percent_identity) if blast else "",
                    "post_trim_query_coverage": format_float(blast.query_coverage) if blast else "",
                    "post_trim_aligned_bases": blast.aligned_bases if blast else "",
                    "post_trim_evalue": format_float(blast.evalue) if blast else "",
                    "post_trim_bitscore": format_float(blast.bitscore) if blast else "",
                    "trimmed_sequence": row.trimmed_sequence,
                }
            )


def write_pivot_csv(rows: list[NormalizedRow], samples: list[str], path: Path) -> None:
    grouped: dict[str, list[NormalizedRow]] = defaultdict(list)
    for row in rows:
        grouped[row.trimmed_sequence_id].append(row)
    fields = [
        "trimmed_sequence_id",
        "occurrence_count",
        "sample_count",
        "total_cluster_reads",
        "trimmed_length",
        "mapping_closest_match_id",
        "mapping_closest_reference",
        "mapping_match_class",
        "mapping_nucleotides_different",
        "mapping_score",
        "post_trim_blast_reference",
        "post_trim_percent_identity",
        "post_trim_query_coverage",
    ] + samples
    ordered = sorted(
        grouped.items(),
        key=lambda item: (
            -len({row.original.sample for row in item[1]}),
            -sum(row.original.cluster_reads for row in item[1]),
            item[0],
        ),
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for sequence_id, group in ordered:
            reads_by_sample: dict[str, int] = defaultdict(int)
            for row in group:
                reads_by_sample[row.original.sample] += row.original.cluster_reads
            representative = sorted(group, key=lambda row: (-row.original.cluster_reads, row.original.sample))[0]
            blast = representative.blast_hit
            writer.writerow(
                {
                    "trimmed_sequence_id": sequence_id,
                    "occurrence_count": len(group),
                    "sample_count": len(reads_by_sample),
                    "total_cluster_reads": sum(row.original.cluster_reads for row in group),
                    "trimmed_length": representative.trimmed_length,
                    "mapping_closest_match_id": representative.mapping_closest_match_id,
                    "mapping_closest_reference": representative.minimap_allele,
                    "mapping_match_class": representative.mapping_match_class,
                    "mapping_nucleotides_different": "" if representative.mapping_nucleotides_different is None else representative.mapping_nucleotides_different,
                    "mapping_score": "" if representative.minimap_score is None else representative.minimap_score,
                    "post_trim_blast_reference": blast.closest_reference if blast else "",
                    "post_trim_percent_identity": format_float(blast.percent_identity) if blast else "",
                    "post_trim_query_coverage": format_float(blast.query_coverage) if blast else "",
                    **{sample: reads_by_sample.get(sample, "") for sample in samples},
                }
            )


def write_trimmed_fasta(rows: list[NormalizedRow], path: Path) -> None:
    records = [
        (
            f"{row.original.sample}_{row.original.cluster}_{row.trimmed_sequence_id}_ReadCount-{row.original.cluster_reads}",
            row.trimmed_sequence,
        )
        for row in sorted(rows, key=lambda item: (item.original.sample, item.original.cluster))
    ]
    write_fasta(records, path)


def write_grouped_trimmed_fasta(rows: list[NormalizedRow], path: Path) -> None:
    grouped: dict[str, list[NormalizedRow]] = defaultdict(list)
    for row in rows:
        grouped[row.trimmed_sequence_id].append(row)
    ordered = sorted(
        grouped.items(),
        key=lambda item: (
            -len({row.original.sample for row in item[1]}),
            -sum(row.original.cluster_reads for row in item[1]),
            item[0],
        ),
    )
    records = []
    for sequence_id, group in ordered:
        representative = sorted(group, key=lambda row: (-row.original.cluster_reads, row.original.sample))[0]
        samples = sorted({row.original.sample for row in group})
        header = (
            f"{sequence_id}"
            f"|occurrences={len(group)}"
            f"|sample_count={len(samples)}"
            f"|samples={';'.join(samples)}"
            f"|total_cluster_reads={sum(row.original.cluster_reads for row in group)}"
        )
        records.append((header, representative.trimmed_sequence))
    write_fasta(records, path)


def format_float(value: float) -> str:
    text = f"{value:.6g}"
    return text


def write_summary(rows: list[NormalizedRow], samples: list[str], path: Path) -> dict[str, object]:
    raw_groups: dict[str, set[str]] = defaultdict(set)
    trimmed_groups: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        raw_groups[row.raw_sequence_id].add(row.original.sample)
        trimmed_groups[row.trimmed_sequence_id].add(row.original.sample)
    summary = {
        "samples": samples,
        "rowCount": len(rows),
        "trimmedRowCount": sum(1 for row in rows if row.trim_start is not None),
        "rawExactSharedGroups": sum(1 for values in raw_groups.values() if len(values) > 1),
        "trimmedExactSharedGroups": sum(1 for values in trimmed_groups.values() if len(values) > 1),
        "rawExactSharedGroupsAcrossAllSelectedSamples": sum(1 for values in raw_groups.values() if values == set(samples)),
        "trimmedExactSharedGroupsAcrossAllSelectedSamples": sum(1 for values in trimmed_groups.values() if values == set(samples)),
    }
    path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def write_provenance(
    args: argparse.Namespace,
    started: dt.datetime,
    wall_time: float,
    steps: list[dict[str, object]],
    inputs: list[Path],
    outputs: list[Path],
    stderr: str,
    exit_status: int,
) -> None:
    output_dir = Path(args.output_dir)
    provenance_path = output_dir / ".lungfish-provenance.json"
    files = [file_record(path, "input", infer_format(path)) for path in unique_existing(inputs)]
    files += [file_record(path, "output", infer_format(path)) for path in unique_existing(outputs)]
    record = {
        "schemaVersion": 1,
        "workflowName": "full-length-ont-mhc-unmatched-normalization-prototype",
        "workflowVersion": TOOL_VERSION,
        "toolName": TOOL_NAME,
        "toolVersion": TOOL_VERSION,
        "createdAt": dt.datetime.now(dt.UTC).isoformat().replace("+00:00", "Z"),
        "startTime": started.isoformat().replace("+00:00", "Z"),
        "endTime": dt.datetime.now(dt.UTC).isoformat().replace("+00:00", "Z"),
        "argv": sys.argv,
        "reproducibleCommand": shlex.join(sys.argv),
        "reproducibleShellCommand": shlex.join(sys.argv),
        "options": {
            "bundle": str(Path(args.bundle).resolve()),
            "outputDir": str(output_dir.resolve()),
            "samples": args.samples,
            "workbook": str(Path(args.workbook).resolve()) if args.workbook else None,
            "workbookSheet": args.workbook_sheet,
            "mappingReference": str(Path(args.mapping_reference).resolve()) if args.mapping_reference else "from provenance",
            "blastReference": str(Path(args.blast_reference).resolve()) if args.blast_reference else "bundle .full-length-ont-mhc/blast-rescue/reference.fa",
            "threads": args.threads,
            "postTrimBlast": args.post_trim_blast,
            "blastEvalue": args.blast_evalue,
            "defaults": {
                "workbookSheet": DEFAULT_WORKBOOK_SHEET,
                "blastEvalue": DEFAULT_BLAST_EVALUE,
            },
        },
        "runtimeIdentity": {
            "python": sys.version,
            "platform": platform.platform(),
            "executable": sys.executable,
            "cwd": str(Path.cwd()),
            "minimap2": shutil.which(args.minimap2) or args.minimap2,
            "blastn": shutil.which(args.blastn) or args.blastn,
            "makeblastdb": shutil.which(args.makeblastdb) or args.makeblastdb,
        },
        "output": str(output_dir.resolve()),
        "files": files,
        "inputs": [file_record(path, "input", infer_format(path)) for path in unique_existing(inputs)],
        "outputs": [file_record(path, "output", infer_format(path)) for path in unique_existing(outputs)],
        "steps": steps,
        "exitStatus": exit_status,
        "wallTimeSeconds": wall_time,
        "stderr": stderr,
    }
    provenance_path.write_text(json.dumps(record, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def unique_existing(paths: Iterable[Path]) -> list[Path]:
    seen: set[Path] = set()
    result: list[Path] = []
    for path in paths:
        resolved = path.resolve()
        if resolved in seen or not resolved.exists() or not resolved.is_file():
            continue
        seen.add(resolved)
        result.append(resolved)
    return result


def infer_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".fa", ".fasta", ".fna"}:
        return "fasta"
    if suffix == ".sam":
        return "sam"
    if suffix == ".tsv":
        return "tsv"
    if suffix == ".csv":
        return "csv"
    if suffix == ".json":
        return "json"
    if suffix == ".xlsx":
        return "xlsx"
    return suffix.lstrip(".") or "file"


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bundle", required=True, help="Input .lungfishgenotype bundle")
    parser.add_argument("--output-dir", required=True, help="Derived output directory")
    parser.add_argument("--samples", nargs="*", help="Sample IDs to process; default: all rows in workbook sheet")
    parser.add_argument("--workbook", help="Workbook path; default: primary workbook from genotype-result.json")
    parser.add_argument("--workbook-sheet", default=DEFAULT_WORKBOOK_SHEET)
    parser.add_argument("--mapping-reference", help="Reference query FASTA for minimap2; default: first minimap2 reference in provenance")
    parser.add_argument("--blast-reference", help="Reference FASTA for post-trim blastn; default: bundle rescue reference.fa")
    parser.add_argument("--threads", type=int, default=max(1, min(4, os.cpu_count() or 1)))
    parser.add_argument("--minimap2", default="minimap2")
    parser.add_argument("--blastn", default="blastn")
    parser.add_argument("--makeblastdb", default="makeblastdb")
    parser.add_argument("--blast-evalue", default=DEFAULT_BLAST_EVALUE)
    parser.add_argument("--post-trim-blast", action="store_true", help="Run blastn after mapping-boundary trimming; skipped by default")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    started = dt.datetime.now(dt.UTC)
    start_time = time.monotonic()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    steps: list[dict[str, object]] = []
    inputs: list[Path] = []
    outputs: list[Path] = []
    stderr = ""
    exit_status = 1
    try:
        bundle = Path(args.bundle)
        workbook_path = Path(args.workbook) if args.workbook else default_workbook_path(bundle)
        samples = args.samples or None
        sample_set = set(samples) if samples else None
        rows = load_workbook_rows(workbook_path, args.workbook_sheet, sample_set)
        if not rows:
            raise ValueError("no workbook rows selected")
        ordered_samples = samples or sorted({row.sample for row in rows})
        mapping_reference = Path(args.mapping_reference) if args.mapping_reference else mapping_reference_from_provenance(bundle)
        blast_reference = Path(args.blast_reference) if args.blast_reference else bundle / ".full-length-ont-mhc" / "blast-rescue" / "reference.fa"
        if not blast_reference.is_file():
            raise FileNotFoundError(f"missing blast reference: {blast_reference}")
        inputs += [bundle / "genotype-result.json", workbook_path, mapping_reference, blast_reference]

        minimap_hits, minimap_steps, minimap_inputs = run_minimap_for_samples(
            bundle=bundle,
            samples=ordered_samples,
            mapping_reference=mapping_reference,
            output_dir=output_dir,
            threads=args.threads,
            minimap2=args.minimap2,
        )
        steps += minimap_steps
        inputs += minimap_inputs

        normalized = trim_rows(rows, minimap_hits)
        if args.post_trim_blast:
            normalized, blast_steps, query_fasta, blast_tsv = run_blast(
                normalized=normalized,
                blast_reference=blast_reference,
                output_dir=output_dir,
                blastn=args.blastn,
                makeblastdb=args.makeblastdb,
                evalue=args.blast_evalue,
                threads=args.threads,
            )
            steps += blast_steps
            inputs.append(query_fasta)
            outputs.append(blast_tsv)

        detail_csv = output_dir / "trimmed-unmatched-detail.csv"
        pivot_csv = output_dir / "trimmed-unmatched-pivot.csv"
        fasta = output_dir / "trimmed-unmatched-clusters.fasta"
        grouped_fasta = output_dir / "trimmed-unmatched-cluster-groups.fasta"
        summary_json = output_dir / "summary.json"
        write_detail_csv(normalized, detail_csv)
        write_pivot_csv(normalized, ordered_samples, pivot_csv)
        write_trimmed_fasta(normalized, fasta)
        write_grouped_trimmed_fasta(normalized, grouped_fasta)
        summary = write_summary(normalized, ordered_samples, summary_json)
        outputs += [detail_csv, pivot_csv, fasta, grouped_fasta, summary_json]
        print(json.dumps(summary, indent=2, sort_keys=True))
        exit_status = 0
        return 0
    except Exception as error:
        stderr = str(error)
        print(stderr, file=sys.stderr)
        return 1
    finally:
        write_provenance(
            args=args,
            started=started,
            wall_time=time.monotonic() - start_time,
            steps=steps,
            inputs=inputs,
            outputs=outputs,
            stderr=stderr,
            exit_status=exit_status,
        )


if __name__ == "__main__":
    raise SystemExit(main())
