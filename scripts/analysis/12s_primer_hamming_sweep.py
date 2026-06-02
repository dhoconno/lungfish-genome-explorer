#!/usr/bin/env python3
"""Sweep 12S primer Hamming thresholds and emit QC summaries.

This is a read-only analysis helper for MIDORI-style source directories with:
  intermediate/12s_reference.fasta
  intermediate/12s_reference.tsv

It intentionally mirrors the existing extraction model: find forward and reverse
primer-like sites in amplicon order, extract the insert between them, and keep
the shortest valid insert for each threshold pair.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import platform
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable


PRIMER_F = "ACTGGGATTAGATACCCC"
PRIMER_R = "CTAGAGGAGCCTGTTCTA"
COMP = str.maketrans("ACGTNacgtn", "TGCANtgcan")


@dataclass(frozen=True)
class Candidate:
    amp: str
    length: int
    f_mm: int
    r_mm: int
    strand: str
    f_start: int
    r_start: int
    f_window: str
    r_window: str

    @property
    def mm_sum(self) -> int:
        return self.f_mm + self.r_mm


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source-dir",
        required=True,
        type=Path,
        help="Source directory containing intermediate/12s_reference.fasta and .tsv",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        type=Path,
        help="Directory where sweep TSV/JSON outputs will be written",
    )
    parser.add_argument("--max-mismatches", type=int, default=6)
    parser.add_argument("--min-insert-length", type=int, default=40)
    parser.add_argument("--max-insert-length", type=int, default=160)
    parser.add_argument(
        "--baseline-forward",
        type=int,
        default=2,
        help="Forward-primer threshold used as baseline for delta columns",
    )
    parser.add_argument(
        "--baseline-reverse",
        type=int,
        default=2,
        help="Reverse-primer threshold used as baseline for delta columns",
    )
    parser.add_argument(
        "--previous-workbook",
        type=Path,
        default=None,
        help="Optional workbook containing a sentinel row to track, e.g. goatfish",
    )
    parser.add_argument(
        "--sentinel-label",
        default="Thick-lipped Goatfish",
        help="Label to search in workbook common-name/taxonomy fields",
    )
    parser.add_argument(
        "--sentinel-sequence",
        default=None,
        help="Optional sentinel insert sequence to track if no workbook is provided",
    )
    parser.add_argument("--max-examples", type=int, default=250)
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_record(path: Path, role: str) -> dict[str, object]:
    return {
        "path": str(path),
        "role": role,
        "sizeBytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def reverse_complement(sequence: str) -> str:
    return sequence.translate(COMP)[::-1]


def iter_fasta(path: Path) -> Iterable[tuple[str, str]]:
    current_id: str | None = None
    chunks: list[str] = []
    with path.open() as handle:
        for raw in handle:
            line = raw.strip()
            if not line:
                continue
            if line.startswith(">"):
                if current_id is not None:
                    yield current_id, "".join(chunks)
                current_id = line[1:].strip()
                chunks = []
            else:
                chunks.append(line)
    if current_id is not None:
        yield current_id, "".join(chunks)


def hamming_hits(sequence: str, primer: str, max_mm: int) -> list[tuple[int, int, str]]:
    length = len(primer)
    hits: list[tuple[int, int, str]] = []
    for start in range(0, len(sequence) - length + 1):
        mm = 0
        for offset, expected in enumerate(primer):
            if sequence[start + offset] != expected:
                mm += 1
                if mm > max_mm:
                    break
        if mm <= max_mm:
            hits.append((start, mm, sequence[start : start + length]))
    return hits


def record_candidates(
    sequence: str,
    max_mm: int,
    min_insert_length: int,
    max_insert_length: int,
) -> list[Candidate]:
    candidates: list[Candidate] = []
    for strand_sequence, strand in (
        (sequence.upper(), "+"),
        (reverse_complement(sequence.upper()), "-"),
    ):
        f_hits = hamming_hits(strand_sequence, PRIMER_F, max_mm)
        if not f_hits:
            continue
        r_hits = hamming_hits(strand_sequence, PRIMER_R, max_mm)
        if not r_hits:
            continue
        for f_start, f_mm, f_window in f_hits:
            insert_start = f_start + len(PRIMER_F)
            for r_start, r_mm, r_window in r_hits:
                if r_start <= insert_start:
                    continue
                insert = strand_sequence[insert_start:r_start]
                insert_length = len(insert)
                if min_insert_length <= insert_length <= max_insert_length:
                    candidates.append(
                        Candidate(
                            amp=insert,
                            length=insert_length,
                            f_mm=f_mm,
                            r_mm=r_mm,
                            strand=strand,
                            f_start=f_start,
                            r_start=r_start,
                            f_window=f_window,
                            r_window=r_window,
                        )
                    )
    return candidates


def selected_candidate(
    candidates: list[Candidate],
    forward_threshold: int,
    reverse_threshold: int,
) -> tuple[Candidate | None, list[Candidate]]:
    eligible = [
        candidate
        for candidate in candidates
        if candidate.f_mm <= forward_threshold and candidate.r_mm <= reverse_threshold
    ]
    if not eligible:
        return None, []
    # Existing in-silico extraction uses shortest valid insert. Tie-breakers make
    # this analysis deterministic without changing the primary criterion.
    return (
        min(
            eligible,
            key=lambda c: (
                c.length,
                c.mm_sum,
                c.f_mm,
                c.r_mm,
                c.strand,
                c.f_start,
                c.r_start,
                c.amp,
            ),
        ),
        eligible,
    )


def load_metadata(path: Path) -> dict[str, dict[str, str]]:
    with path.open() as handle:
        return {row["seq_id"]: row for row in csv.DictReader(handle, delimiter="\t")}


def sentinel_from_workbook(path: Path, label: str) -> str | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        common_name = str(values.get("comon name") or values.get("common name") or "")
        taxonomy = str(values.get("taxonomy") or "")
        if label.lower() in common_name.lower() or label.lower() in taxonomy.lower():
            sequence = values.get("seq")
            if sequence:
                return str(sequence).strip().upper()
    return None


def write_tsv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, object]]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main() -> int:
    args = parse_args()
    start = time.time()
    start_iso = datetime.now(timezone.utc).isoformat()
    source_dir = args.source_dir
    fasta_path = source_dir / "intermediate" / "12s_reference.fasta"
    metadata_path = source_dir / "intermediate" / "12s_reference.tsv"
    if not fasta_path.exists():
        raise FileNotFoundError(fasta_path)
    if not metadata_path.exists():
        raise FileNotFoundError(metadata_path)

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    sentinel_sequence = args.sentinel_sequence
    if sentinel_sequence is None and args.previous_workbook:
        sentinel_sequence = sentinel_from_workbook(args.previous_workbook, args.sentinel_label)
    sentinel_sequence = sentinel_sequence.upper() if sentinel_sequence else None
    sentinel_sha16 = (
        hashlib.sha256(sentinel_sequence.encode()).hexdigest()[:16]
        if sentinel_sequence
        else ""
    )

    metadata = load_metadata(metadata_path)
    thresholds = range(args.max_mismatches + 1)
    configs = [(forward, reverse) for forward in thresholds for reverse in thresholds]
    baseline_key = (args.baseline_forward, args.baseline_reverse)

    stats: dict[tuple[int, int], dict[str, object]] = {}
    hit_ids: dict[tuple[int, int], set[str]] = {}
    seqs: dict[tuple[int, int], set[str]] = {}
    amp_groups: dict[tuple[int, int], dict[str, set[str]]] = {}
    examples_new_vs_baseline: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    suspicious_examples: list[dict[str, object]] = []
    sentinel_rows: list[dict[str, object]] = []

    for config in configs:
        stats[config] = {
            "record_hits": 0,
            "group_counts": Counter(),
            "length_counts": Counter(),
            "mismatch_pair_counts": Counter(),
            "multi_candidate_records": 0,
            "multi_insert_records": 0,
            "selected_not_lowest_mismatch_records": 0,
            "short_lt80": 0,
            "long_gt130": 0,
            "outside_90_120": 0,
            "sentinel_records": [],
        }
        hit_ids[config] = set()
        seqs[config] = set()
        amp_groups[config] = defaultdict(set)

    baseline_selected: dict[str, str] = {}
    record_count = 0
    no_candidate_count = 0
    max_candidate_count = 0
    high_candidate_records = 0
    progress_next = 25000

    for seq_id, raw_sequence in iter_fasta(fasta_path):
        record_count += 1
        candidates = record_candidates(
            raw_sequence,
            args.max_mismatches,
            args.min_insert_length,
            args.max_insert_length,
        )
        max_candidate_count = max(max_candidate_count, len(candidates))
        if len(candidates) >= 20:
            high_candidate_records += 1
        if not candidates:
            no_candidate_count += 1
        meta = metadata.get(seq_id, {})
        group = meta.get("group") or "Unknown"

        selected_by_config: dict[tuple[int, int], tuple[Candidate, list[Candidate]]] = {}
        for config in configs:
            selected, eligible = selected_candidate(candidates, config[0], config[1])
            if selected is None:
                continue
            selected_by_config[config] = (selected, eligible)

            config_stats = stats[config]
            config_stats["record_hits"] = int(config_stats["record_hits"]) + 1
            config_stats["group_counts"][group] += 1
            config_stats["length_counts"][selected.length] += 1
            config_stats["mismatch_pair_counts"][(selected.f_mm, selected.r_mm)] += 1
            if selected.length < 80:
                config_stats["short_lt80"] = int(config_stats["short_lt80"]) + 1
            if selected.length > 130:
                config_stats["long_gt130"] = int(config_stats["long_gt130"]) + 1
            if selected.length < 90 or selected.length > 120:
                config_stats["outside_90_120"] = int(config_stats["outside_90_120"]) + 1
            distinct_inserts = {candidate.amp for candidate in eligible}
            if len(eligible) > 1:
                config_stats["multi_candidate_records"] = (
                    int(config_stats["multi_candidate_records"]) + 1
                )
            if len(distinct_inserts) > 1:
                config_stats["multi_insert_records"] = int(config_stats["multi_insert_records"]) + 1
            min_mm_sum = min(candidate.mm_sum for candidate in eligible)
            if selected.mm_sum > min_mm_sum:
                config_stats["selected_not_lowest_mismatch_records"] = (
                    int(config_stats["selected_not_lowest_mismatch_records"]) + 1
                )
                if len(suspicious_examples) < args.max_examples:
                    best_mm_candidate = min(
                        eligible,
                        key=lambda c: (c.mm_sum, abs(c.length - 106), c.length),
                    )
                    suspicious_examples.append(
                        {
                            "forward_threshold": config[0],
                            "reverse_threshold": config[1],
                            "seq_id": seq_id,
                            "common_name": meta.get("common_name", ""),
                            "latin_name": meta.get("latin_name", ""),
                            "group": group,
                            "selected_length": selected.length,
                            "selected_f_mm": selected.f_mm,
                            "selected_r_mm": selected.r_mm,
                            "selected_sha16": hashlib.sha256(selected.amp.encode()).hexdigest()[
                                :16
                            ],
                            "best_mm_length": best_mm_candidate.length,
                            "best_mm_f_mm": best_mm_candidate.f_mm,
                            "best_mm_r_mm": best_mm_candidate.r_mm,
                            "best_mm_sha16": hashlib.sha256(
                                best_mm_candidate.amp.encode()
                            ).hexdigest()[:16],
                            "eligible_candidate_count": len(eligible),
                            "distinct_insert_count": len(distinct_inserts),
                        }
                    )

            hit_ids[config].add(seq_id)
            seqs[config].add(selected.amp)
            amp_groups[config][selected.amp].add(group)
            if sentinel_sequence and selected.amp == sentinel_sequence:
                config_stats["sentinel_records"].append(seq_id)

        if baseline_key in selected_by_config:
            baseline_selected[seq_id] = selected_by_config[baseline_key][0].amp

        for config, (selected, _eligible) in selected_by_config.items():
            if config == baseline_key:
                continue
            if seq_id not in baseline_selected and len(examples_new_vs_baseline[config]) < 50:
                examples_new_vs_baseline[config].append(
                    {
                        "forward_threshold": config[0],
                        "reverse_threshold": config[1],
                        "seq_id": seq_id,
                        "common_name": meta.get("common_name", ""),
                        "latin_name": meta.get("latin_name", ""),
                        "group": group,
                        "insert_length": selected.length,
                        "f_mm": selected.f_mm,
                        "r_mm": selected.r_mm,
                        "f_window": selected.f_window,
                        "r_window": selected.r_window,
                        "insert_sha16": hashlib.sha256(selected.amp.encode()).hexdigest()[:16],
                    }
                )

        if sentinel_sequence:
            for config, (selected, eligible) in selected_by_config.items():
                if selected.amp == sentinel_sequence:
                    sentinel_rows.append(
                        {
                            "forward_threshold": config[0],
                            "reverse_threshold": config[1],
                            "seq_id": seq_id,
                            "common_name": meta.get("common_name", ""),
                            "latin_name": meta.get("latin_name", ""),
                            "group": group,
                            "insert_length": selected.length,
                            "f_mm": selected.f_mm,
                            "r_mm": selected.r_mm,
                            "f_window": selected.f_window,
                            "r_window": selected.r_window,
                            "eligible_candidate_count": len(eligible),
                            "sentinel_sha16": sentinel_sha16,
                        }
                    )

        if record_count >= progress_next:
            elapsed = time.time() - start
            print(f"processed {record_count} records in {elapsed:.1f}s", flush=True)
            progress_next += 25000

    baseline_ids = hit_ids[baseline_key]
    baseline_seqs = seqs[baseline_key]

    summary_rows: list[dict[str, object]] = []
    for config in configs:
        config_stats = stats[config]
        record_hits = int(config_stats["record_hits"])
        unique_count = len(seqs[config])
        new_ids = hit_ids[config] - baseline_ids
        new_seqs = seqs[config] - baseline_seqs
        multi_group_amplicons = sum(1 for groups in amp_groups[config].values() if len(groups) > 1)
        summary_rows.append(
            {
                "forward_threshold": config[0],
                "reverse_threshold": config[1],
                "record_hits": record_hits,
                "new_records_vs_baseline": len(new_ids),
                "unique_amplicons": unique_count,
                "new_unique_amplicons_vs_baseline": len(new_seqs),
                "multi_candidate_records": int(config_stats["multi_candidate_records"]),
                "multi_insert_records": int(config_stats["multi_insert_records"]),
                "selected_not_lowest_mismatch_records": int(
                    config_stats["selected_not_lowest_mismatch_records"]
                ),
                "short_lt80": int(config_stats["short_lt80"]),
                "long_gt130": int(config_stats["long_gt130"]),
                "outside_90_120": int(config_stats["outside_90_120"]),
                "multi_group_amplicons": multi_group_amplicons,
                "sentinel_record_count": len(config_stats["sentinel_records"]),
                "sentinel_records": ",".join(config_stats["sentinel_records"]),
            }
        )

    write_tsv(
        output_dir / "summary.tsv",
        [
            "forward_threshold",
            "reverse_threshold",
            "record_hits",
            "new_records_vs_baseline",
            "unique_amplicons",
            "new_unique_amplicons_vs_baseline",
            "multi_candidate_records",
            "multi_insert_records",
            "selected_not_lowest_mismatch_records",
            "short_lt80",
            "long_gt130",
            "outside_90_120",
            "multi_group_amplicons",
            "sentinel_record_count",
            "sentinel_records",
        ],
        summary_rows,
    )

    matrix_rows = []
    for forward in thresholds:
        row = {"forward_threshold": forward}
        for reverse in thresholds:
            row[f"reverse_{reverse}"] = int(stats[(forward, reverse)]["record_hits"])
        matrix_rows.append(row)
    write_tsv(
        output_dir / "record_hit_matrix.tsv",
        ["forward_threshold"] + [f"reverse_{reverse}" for reverse in thresholds],
        matrix_rows,
    )

    length_rows = []
    mismatch_rows = []
    new_group_rows = []
    multi_group_rows = []
    for config in configs:
        for length, count in sorted(stats[config]["length_counts"].items()):
            length_rows.append(
                {
                    "forward_threshold": config[0],
                    "reverse_threshold": config[1],
                    "insert_length": length,
                    "count": count,
                }
            )
        for pair, count in sorted(stats[config]["mismatch_pair_counts"].items()):
            mismatch_rows.append(
                {
                    "forward_threshold": config[0],
                    "reverse_threshold": config[1],
                    "f_mm": pair[0],
                    "r_mm": pair[1],
                    "count": count,
                }
            )
        new_ids = hit_ids[config] - baseline_ids
        group_counts = Counter(metadata.get(seq_id, {}).get("group") or "Unknown" for seq_id in new_ids)
        for group, count in group_counts.most_common():
            new_group_rows.append(
                {
                    "forward_threshold": config[0],
                    "reverse_threshold": config[1],
                    "group": group,
                    "new_records_vs_baseline": count,
                }
            )
        emitted = 0
        for amp, groups in amp_groups[config].items():
            if len(groups) <= 1:
                continue
            multi_group_rows.append(
                {
                    "forward_threshold": config[0],
                    "reverse_threshold": config[1],
                    "insert_sha16": hashlib.sha256(amp.encode()).hexdigest()[:16],
                    "insert_length": len(amp),
                    "groups": ",".join(sorted(groups)),
                }
            )
            emitted += 1
            if emitted >= 100:
                break

    write_tsv(
        output_dir / "length_histogram.tsv",
        ["forward_threshold", "reverse_threshold", "insert_length", "count"],
        length_rows,
    )
    write_tsv(
        output_dir / "selected_mismatch_pairs.tsv",
        ["forward_threshold", "reverse_threshold", "f_mm", "r_mm", "count"],
        mismatch_rows,
    )
    write_tsv(
        output_dir / "new_records_by_group_vs_baseline.tsv",
        ["forward_threshold", "reverse_threshold", "group", "new_records_vs_baseline"],
        new_group_rows,
    )
    write_tsv(
        output_dir / "multi_group_amplicon_examples.tsv",
        ["forward_threshold", "reverse_threshold", "insert_sha16", "insert_length", "groups"],
        multi_group_rows,
    )

    example_rows = [
        row
        for config in configs
        for row in examples_new_vs_baseline.get(config, [])
        if config[0] >= args.baseline_forward or config[1] >= args.baseline_reverse
    ]
    write_tsv(
        output_dir / "new_hit_examples_vs_baseline.tsv",
        [
            "forward_threshold",
            "reverse_threshold",
            "seq_id",
            "common_name",
            "latin_name",
            "group",
            "insert_length",
            "f_mm",
            "r_mm",
            "f_window",
            "r_window",
            "insert_sha16",
        ],
        example_rows,
    )
    write_tsv(
        output_dir / "suspicious_selection_examples.tsv",
        [
            "forward_threshold",
            "reverse_threshold",
            "seq_id",
            "common_name",
            "latin_name",
            "group",
            "selected_length",
            "selected_f_mm",
            "selected_r_mm",
            "selected_sha16",
            "best_mm_length",
            "best_mm_f_mm",
            "best_mm_r_mm",
            "best_mm_sha16",
            "eligible_candidate_count",
            "distinct_insert_count",
        ],
        suspicious_examples,
    )
    write_tsv(
        output_dir / "sentinel_recovery.tsv",
        [
            "forward_threshold",
            "reverse_threshold",
            "seq_id",
            "common_name",
            "latin_name",
            "group",
            "insert_length",
            "f_mm",
            "r_mm",
            "f_window",
            "r_window",
            "eligible_candidate_count",
            "sentinel_sha16",
        ],
        sentinel_rows,
    )

    end_iso = datetime.now(timezone.utc).isoformat()
    output_files = sorted(path for path in output_dir.iterdir() if path.is_file())
    provenance = {
        "schemaVersion": 1,
        "workflowName": "12s-primer-hamming-sweep",
        "workflowVersion": "1",
        "toolName": "scripts/analysis/12s_primer_hamming_sweep.py",
        "argv": sys.argv,
        "reproducibleCommand": " ".join(sys.argv),
        "startTime": start_iso,
        "endTime": end_iso,
        "wallTimeSeconds": time.time() - start,
        "status": "success",
        "exitStatus": 0,
        "runtime": {
            "python": sys.version,
            "executable": sys.executable,
            "platform": platform.platform(),
        },
        "parameters": {
            "sourceDir": str(source_dir),
            "maxMismatches": args.max_mismatches,
            "minInsertLength": args.min_insert_length,
            "maxInsertLength": args.max_insert_length,
            "baselineForward": args.baseline_forward,
            "baselineReverse": args.baseline_reverse,
            "primerF": PRIMER_F,
            "primerR": PRIMER_R,
            "previousWorkbook": str(args.previous_workbook) if args.previous_workbook else None,
            "sentinelLabel": args.sentinel_label,
            "sentinelSha16": sentinel_sha16,
        },
        "inputs": [file_record(fasta_path, "input"), file_record(metadata_path, "input")]
        + ([file_record(args.previous_workbook, "input")] if args.previous_workbook else []),
        "outputs": [file_record(path, "output") for path in output_files if path.name != "provenance.json"],
        "summary": {
            "recordCount": record_count,
            "recordsWithoutAnyCandidateAtMaxThreshold": no_candidate_count,
            "maxCandidateCountPerRecord": max_candidate_count,
            "recordsWithAtLeast20Candidates": high_candidate_records,
            "baseline": f"{args.baseline_forward}/{args.baseline_reverse}",
        },
    }
    provenance_path = output_dir / "provenance.json"
    provenance_path.write_text(json.dumps(provenance, indent=2, sort_keys=True) + "\n")

    print(f"wrote {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
