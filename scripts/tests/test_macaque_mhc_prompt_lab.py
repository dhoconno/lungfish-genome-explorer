import json
import shlex
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.analysis import macaque_mhc_prompt_lab as lab


class MacaqueMHCPromptLabTests(unittest.TestCase):
    def make_synthetic_snprc_workbook(self, root: Path) -> Path:
        path = root / "synthetic_snprc.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Full Sequencing Results 1"
        ws.append(["Client ID", None, None, "44470", "44395"])
        ws.append(["GS ID", "Total", "Average", "LC1729", "LC1730"])
        ws.append(["Mapped Read Count", None, None, 1000, 1200])
        ws.append(["total_read_count", None, None, 2000, 2200])
        ws.append(["percent_reads_unmapped", None, None, 50.0, 45.4])
        ws.append(["MHC-A Haplotype 1", None, None, "A002.01", "A008.01"])
        ws.append(["MHC-A Haplotype 2", None, None, "A002.01", "A004.01"])
        ws.append(["MHC-B Haplotype 1", None, None, "B001.01", "B028.01"])
        ws.append(["MHC-B Haplotype 2", None, None, "B001.01", "B012.01"])
        ws.append(["MHC-DRB Haplotype 1", None, None, "DR09.01", "DR06.01"])
        ws.append(["MHC-DRB Haplotype 2", None, None, "DR09.01", "DR01.01"])
        ws.append(["MHC-DQA Haplotype 1", None, None, "26g2", "23_01"])
        ws.append(["MHC-DQA Haplotype 2", None, None, "26g2", "01g2"])
        ws.append(["MHC-DQB Haplotype 1", None, None, "18g3", "06g1"])
        ws.append(["MHC-DQB Haplotype 2", None, None, "18g3", "06g2"])
        ws.append(["MHC-DPA Haplotype 1", None, None, "02g1", "02g3"])
        ws.append(["MHC-DPA Haplotype 2", None, None, "02g1", "02g1"])
        ws.append(["MHC-DPB Haplotype 1", None, None, "15g", "07g1"])
        ws.append(["MHC-DPB Haplotype 2", None, None, "15g", "08_01"])
        ws.append(["Comments", "Subtotal", "# Obs.", None, None])
        ws.append(["Mamu-A Major Alleles", None, None, None, None])
        ws.append(["01_Mamu-A1_002g", 100, 1, 90, None])
        ws.append(["01_Mamu-A1_008g", 100, 1, None, 80])
        ws.append(["01_Mamu-A1_004g", 100, 1, None, 70])
        ws.append(["Mamu-B Major Alleles", None, None, None, None])
        ws.append(["03_Mamu-B_001g1", 100, 1, 75, None])
        ws.append(["03_Mamu-B_028_01_01_01", 100, 1, None, 60])
        ws.append(["03_Mamu-B_012g", 100, 1, None, 55])
        ws.append(["Mamu-DRB1 Alleles", None, None, None, None])
        ws.append(["07_Mamu-DRB1_03_03_01_01", 100, 1, 65, None])
        ws.append(["07_Mamu-DRB1_03_12_01_01", 100, 1, None, 45])
        ws.append(["Mamu-DQA Alleles", None, None, None, None])
        ws.append(["09_Mamu-DQA1_26_01", 100, 1, 100, None])
        ws.append(["09_Mamu-DQA1_23_02", 100, 1, None, 95])
        ws.append(["Mamu-DQB Alleles", None, None, None, None])
        ws.append(["10_Mamu-DQB1_18g3", 100, 1, 110, None])
        ws.append(["10_Mamu-DQB1_06g1", 100, 1, None, 105])
        ws.append(["Mamu-DPA Alleles", None, None, None, None])
        ws.append(["11_Mamu-DPA1_02g1", 100, 1, 120, 30])
        ws.append(["11_Mamu-DPA1_02g3", 100, 1, None, 90])
        ws.append(["Mamu-DPB Alleles", None, None, None, None])
        ws.append(["12_Mamu-DPB1_15g", 100, 1, 125, None])
        ws.append(["12_Mamu-DPB1_07g1", 100, 1, None, 98])
        wb.create_sheet("Full Sequencing Results 2")
        wb.save(path)
        return path

    def append_truth_rows(self, ws, sample_values: list[tuple[str, str]]) -> None:
        for locus in lab.REPORT_LOCI:
            for slot in (1, 2):
                row = [f"{locus} Haplotype {slot}", None, None]
                row.extend(value for pair in sample_values for value in pair[slot - 1:slot])
                ws.append(row)

    def test_extract_blinds_truth_and_preserves_read_counts(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self.make_synthetic_snprc_workbook(Path(temp))
            extracted = lab.extract_workbook(workbook)

        self.assertEqual([sample["gs_id"] for sample in extracted["samples"]], ["LC1729", "LC1730"])
        truth = extracted["truth_calls"]
        self.assertEqual(truth["LC1729"]["MHC-A"], ["A002.01", "A002.01"])
        self.assertEqual(truth["LC1730"]["MHC-DPB"], ["07g1", "08_01"])
        observations = extracted["prompt_input"]["observations"]
        self.assertIn(
            {
                "sample_id": "LC1729",
                "client_id": "44470",
                "report_locus": "MHC-A",
                "source_locus": "Mamu-A1",
                "genotype": "01_Mamu-A1_002g",
                "reads": 90,
                "sheet": "Full Sequencing Results 1",
                "row": 22,
                "sample_mapped_reads": 1000,
                "read_fraction": 0.09,
            },
            observations,
        )
        prompt_json = json.dumps(extracted["prompt_input"], sort_keys=True)
        self.assertNotIn("A002.01", prompt_json)
        self.assertNotIn("B001.01", prompt_json)
        self.assertNotIn("DR09.01", prompt_json)

    def test_render_prompt_includes_generalist_rules_and_blinded_input(self):
        prompt_template = "# Prompt\n\n{{PROMPT_INPUT_JSON}}\n"
        prompt_input = {
            "schema_version": 1,
            "dataset": "synthetic.xlsx",
            "instructions": {"truth_blinded": True, "report_loci": lab.REPORT_LOCI},
            "samples": [{"gs_id": "LC1729", "client_id": "44470"}],
            "observations": [{"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}],
        }
        rendered = lab.render_prompt_text(prompt_template, prompt_input)
        self.assertIn('"truth_blinded": true', rendered)
        self.assertIn("01_Mamu-A1_002g", rendered)
        self.assertNotIn("A002.01", rendered)

    def test_committed_prompt_template_defines_concrete_output_schema(self):
        template = lab.DEFAULT_PROMPT.read_text(encoding="utf-8")
        haplotype_schema = template.split("haplotype_definitions must be a list of objects with:", 1)[1].split(
            "sample_calls must be a list of objects with:",
            1,
        )[0]
        unresolved_schema = template.split("unresolved must be a list of objects", 1)[1].split(
            "Allowed confidence values",
            1,
        )[0]

        self.assertEqual(template.count("{{PROMPT_INPUT_JSON}}"), 1)
        self.assertIn("schema_version must be integer 1", template)
        self.assertIn("- label", haplotype_schema)
        self.assertNotIn("- haplotype", haplotype_schema)
        self.assertIn("- reason", unresolved_schema)
        self.assertIn("- evidence_summary", unresolved_schema)
        template_without_allowed_prohibition = template.replace("Do not use any MCM M1-M7 prior", "").replace(
            "do not use any MCM M1-M7 prior",
            "",
        )
        self.assertNotRegex(template_without_allowed_prohibition, r"\bM[1-7]\b")
        self.assertNotIn("M1-M7", template_without_allowed_prohibition)

    def test_render_prompt_rejects_missing_prompt_input_placeholder(self):
        with self.assertRaisesRegex(ValueError, "exactly one"):
            lab.render_prompt_text("# Prompt\n", {"schema_version": 1})

    def test_render_prompt_rejects_duplicate_prompt_input_placeholders(self):
        prompt_template = "# Prompt\n\n{{PROMPT_INPUT_JSON}}\n\n{{PROMPT_INPUT_JSON}}\n"
        with self.assertRaisesRegex(ValueError, "exactly one"):
            lab.render_prompt_text(prompt_template, {"schema_version": 1})

    def test_locus_mapping_keeps_report_loci_separate(self):
        cases = {
            "01_Mamu-A1_002g": ("MHC-A", "Mamu-A1"),
            "15_Mamu-AG3_02_A1_028": ("MHC-A", "Mamu-AG3"),
            "03_Mamu-B_001g1": ("MHC-B", "Mamu-B"),
            "06_Mamu-I_01g1": ("MHC-B", "Mamu-I"),
            "07_Mamu-DRB1_03_12_01_01": ("MHC-DRB", "Mamu-DRB1"),
            "09_Mamu-DQA1_26_01": ("MHC-DQA", "Mamu-DQA1"),
            "10_Mamu-DQB1_18g3": ("MHC-DQB", "Mamu-DQB1"),
            "11_Mamu-DPA1_02g1": ("MHC-DPA", "Mamu-DPA1"),
            "12_Mamu-DPB1_15g": ("MHC-DPB", "Mamu-DPB1"),
            "13_Mamu-E_02g1": ("context", "Mamu-E"),
        }
        for genotype, expected in cases.items():
            with self.subTest(genotype=genotype):
                self.assertEqual(lab.locus_from_genotype(genotype), expected)

    def test_validate_model_output_accepts_expected_shape(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        output = {
            "schema_version": 1,
            "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
            "haplotype_definitions": [
                {
                    "locus": "MHC-A",
                    "label": "A-A1*002-H01",
                    "supporting_genotypes": ["01_Mamu-A1_002g"],
                    "seed_samples": ["LC1729"],
                    "confidence": "high",
                    "rationale": "seed",
                }
            ],
            "sample_calls": [
                {
                    "sample_id": "LC1729",
                    "locus": "MHC-A",
                    "h1": "A-A1*002-H01",
                    "h2": "A-A1*002-H01",
                    "status": "called",
                    "h1_supporting_genotypes": ["01_Mamu-A1_002g"],
                    "h2_supporting_genotypes": ["01_Mamu-A1_002g"],
                    "rationale": "called",
                }
            ],
            "unresolved": [],
        }
        self.assertEqual(lab.validate_model_output(output, prompt_input), [])

    def test_validate_model_output_rejects_non_integer_schema_version(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        for schema_version in (True, 1.0):
            with self.subTest(schema_version=schema_version):
                output = {
                    "schema_version": schema_version,
                    "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                    "haplotype_definitions": [],
                    "sample_calls": [],
                    "unresolved": [],
                }

                errors = lab.validate_model_output(output, prompt_input)

                self.assertTrue(any("schema_version" in error for error in errors))

    def test_validate_model_output_rejects_unknown_sample_locus_and_genotype(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        output = {
            "schema_version": 1,
            "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
            "haplotype_definitions": [
                {
                    "locus": "MHC-Z",
                    "label": "bad",
                    "supporting_genotypes": ["missing"],
                    "seed_samples": ["LC9999"],
                    "confidence": "high",
                    "rationale": "bad",
                }
            ],
            "sample_calls": [
                {
                    "sample_id": "LC9999",
                    "locus": "MHC-Z",
                    "h1": "bad",
                    "h2": "?",
                    "status": "called",
                    "h1_supporting_genotypes": ["missing"],
                    "h2_supporting_genotypes": [],
                    "rationale": "bad",
                }
            ],
            "unresolved": [],
        }
        errors = lab.validate_model_output(output, prompt_input)
        self.assertTrue(any("unknown sample" in error for error in errors))
        self.assertTrue(any("unknown locus" in error for error in errors))
        self.assertTrue(any("unknown genotype" in error for error in errors))

    def test_validate_model_output_rejects_extra_top_level_keys(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        output = {
            "schema_version": 1,
            "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
            "haplotype_definitions": [],
            "sample_calls": [],
            "unresolved": [],
            "notes": "not part of the exact schema",
        }

        errors = lab.validate_model_output(output, prompt_input)

        self.assertTrue(any("extra top-level key" in error and "notes" in error for error in errors))

    def test_validate_model_output_rejects_malformed_field_types_without_raising(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        output = {
            "schema_version": 1,
            "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
            "haplotype_definitions": [
                {
                    "locus": "MHC-A",
                    "label": "A-A1*002-H01",
                    "supporting_genotypes": [["01_Mamu-A1_002g"]],
                    "seed_samples": [["LC1729"]],
                    "confidence": [],
                    "rationale": "seed",
                }
            ],
            "sample_calls": [
                {
                    "sample_id": ["LC1729"],
                    "locus": ["MHC-A"],
                    "h1": [],
                    "h2": "?",
                    "status": [],
                    "h1_supporting_genotypes": [["01_Mamu-A1_002g"]],
                    "h2_supporting_genotypes": [],
                    "rationale": [],
                }
            ],
            "unresolved": [{"sample_id": ["LC1729"], "locus": ["MHC-A"], "reason": [], "evidence_summary": []}],
        }

        errors = lab.validate_model_output(output, prompt_input)

        self.assertTrue(errors)
        self.assertTrue(any("confidence" in error for error in errors))
        self.assertTrue(any("supporting_genotypes" in error and "must contain strings" in error for error in errors))
        self.assertTrue(any("sample_id" in error and "must be a string" in error for error in errors))
        self.assertTrue(any("h1" in error and "must be a string" in error for error in errors))
        self.assertTrue(any("rationale" in error and "must be a string" in error for error in errors))
        self.assertTrue(any("evidence_summary" in error and "must be a string" in error for error in errors))

    def test_validate_model_output_rejects_non_string_definition_label_and_rationale(self):
        prompt_input = {
            "samples": [{"gs_id": "LC1729"}],
            "instructions": {"report_loci": ["MHC-A"]},
            "observations": [
                {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
            ],
        }
        output = {
            "schema_version": 1,
            "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
            "haplotype_definitions": [
                {
                    "locus": "MHC-A",
                    "label": 42,
                    "supporting_genotypes": ["01_Mamu-A1_002g"],
                    "seed_samples": ["LC1729"],
                    "confidence": "high",
                    "rationale": [],
                }
            ],
            "sample_calls": [],
            "unresolved": [],
        }

        errors = lab.validate_model_output(output, prompt_input)

        self.assertTrue(any("label" in error and "must be a string" in error for error in errors))
        self.assertTrue(any("rationale" in error and "must be a string" in error for error in errors))

    def test_score_maps_predicted_labels_to_human_labels(self):
        truth = {
            "LC1729": {"MHC-A": ["A002.01", "A002.01"]},
            "LC1730": {"MHC-A": ["A008.01", "A004.01"]},
        }
        output = {
            "sample_calls": [
                {"sample_id": "LC1729", "locus": "MHC-A", "h1": "A-A1*002-H01", "h2": "A-A1*002-H01", "status": "called"},
                {"sample_id": "LC1730", "locus": "MHC-A", "h1": "A-A1*008-H02", "h2": "A-A1*004-H03", "status": "called"},
            ]
        }
        score = lab.score_predictions(output, truth, loci=["MHC-A"])
        self.assertEqual(score["overall"]["slot_concordance"], 1.0)
        self.assertEqual(score["overall"]["pair_concordance"], 1.0)
        self.assertEqual(score["loci"]["MHC-A"]["label_mapping"]["A-A1*002-H01"], "A002.01")

    def test_score_label_mapping_is_slot_order_independent(self):
        truth = {
            "LC1": {"MHC-A": ["A001", "A002"]},
            "LC2": {"MHC-A": ["A001", "A003"]},
        }
        output_in_order = {
            "sample_calls": [
                {"sample_id": "LC1", "locus": "MHC-A", "h1": "pred-A001", "h2": "pred-A002", "status": "called"},
                {"sample_id": "LC2", "locus": "MHC-A", "h1": "pred-A001", "h2": "pred-A003", "status": "called"},
            ]
        }
        output_swapped = {
            "sample_calls": [
                {"sample_id": "LC1", "locus": "MHC-A", "h1": "pred-A002", "h2": "pred-A001", "status": "called"},
                {"sample_id": "LC2", "locus": "MHC-A", "h1": "pred-A001", "h2": "pred-A003", "status": "called"},
            ]
        }

        score_in_order = lab.score_predictions(output_in_order, truth, loci=["MHC-A"])
        score_swapped = lab.score_predictions(output_swapped, truth, loci=["MHC-A"])

        for score in (score_in_order, score_swapped):
            self.assertEqual(score["overall"]["slot_concordance"], 1.0)
            self.assertEqual(score["overall"]["pair_concordance"], 1.0)
            self.assertEqual(score["overall"]["false_merge_count"], 0)
            self.assertEqual(score["overall"]["false_split_count"], 0)
        self.assertEqual(
            set(score_in_order["loci"]["MHC-A"]["label_mapping"].items()),
            set(score_swapped["loci"]["MHC-A"]["label_mapping"].items()),
        )

    def test_score_counts_unresolved_and_false_merge(self):
        truth = {
            "LC1": {"MHC-A": ["A001.01", "A002.01"]},
            "LC2": {"MHC-A": ["A001.01", "A003.01"]},
        }
        output = {
            "sample_calls": [
                {"sample_id": "LC1", "locus": "MHC-A", "h1": "A-shared", "h2": "?", "status": "partial"},
                {"sample_id": "LC2", "locus": "MHC-A", "h1": "A-shared", "h2": "A-shared", "status": "called"},
            ]
        }
        score = lab.score_predictions(output, truth, loci=["MHC-A"])
        self.assertGreater(score["overall"]["unresolved_rate"], 0)
        self.assertGreaterEqual(score["loci"]["MHC-A"]["false_merge_count"], 1)

    def test_max_weight_label_mapping_finds_global_optimum(self):
        weights = {"predA": {"humanX": 10, "humanY": 9}, "predB": {"humanX": 9}}

        mapping = lab.max_weight_label_mapping(weights)

        self.assertEqual(mapping, {"predA": "humanY", "predB": "humanX"})

    def test_score_cli_writes_score_reports_and_complete_provenance(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            truth_path = iteration_dir / "truth_calls.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            truth = {
                "LC1729": {"MHC-A": ["A002.01", "A002.01"]},
                "LC1730": {"MHC-A": ["A008.01", "A004.01"]},
            }
            output = {
                "schema_version": 1,
                "prompt_version": lab.PROMPT_VERSION,
                "haplotype_definitions": [],
                "sample_calls": [
                    {
                        "sample_id": "LC1729",
                        "locus": "MHC-A",
                        "h1": "A-A1*002-H01",
                        "h2": "A-A1*002-H01",
                        "status": "called",
                    },
                    {
                        "sample_id": "LC1730",
                        "locus": "MHC-A",
                        "h1": "A-A1*008-H02",
                        "h2": "A-A1*004-H03",
                        "status": "called",
                    },
                ],
                "unresolved": [],
            }
            truth_path.write_text(json.dumps(truth), encoding="utf-8")
            parsed_path.write_text(json.dumps(output), encoding="utf-8")
            argv = ["score", "--iteration-dir", str(iteration_dir)]

            self.assertEqual(lab.main(argv), 0)

            score_path = iteration_dir / "score.json"
            report_path = iteration_dir / "score.md"
            provenance_path = iteration_dir / "score.provenance.json"
            self.assertTrue(score_path.is_file())
            self.assertTrue(report_path.is_file())
            self.assertTrue(provenance_path.is_file())
            score = json.loads(score_path.read_text(encoding="utf-8"))
            self.assertEqual(score["overall"]["slot_concordance"], 1.0)
            self.assertEqual(score["overall"]["pair_concordance"], 1.0)
            self.assertEqual(score["overall"]["unresolved_rate"], 0.0)
            report = report_path.read_text(encoding="utf-8")
            self.assertIn("slot_concordance", report)
            self.assertIn("MHC-A", report)

            provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
            expected_argv = [str(Path(lab.__file__)), *argv]
            self.assertEqual(provenance["schemaVersion"], 1)
            self.assertEqual(provenance["workflowName"], "score")
            self.assertEqual(provenance["toolName"], lab.TOOL_NAME)
            self.assertEqual(provenance["toolVersion"], lab.TOOL_VERSION)
            self.assertEqual(provenance["argv"], expected_argv)
            self.assertEqual(
                provenance["reproducibleShellCommand"],
                " ".join(shlex.quote(part) for part in [sys.executable, *expected_argv]),
            )
            self.assertEqual(provenance["options"]["iterationDir"], str(iteration_dir.resolve()))
            self.assertEqual(provenance["options"]["defaults"]["workbook"], str(lab.DEFAULT_SNPRC_WORKBOOK))
            self.assertEqual(provenance["options"]["defaults"]["prompt"], str(lab.DEFAULT_PROMPT))
            self.assertEqual(provenance["options"]["defaults"]["outputRoot"], str(lab.DEFAULT_OUTPUT_ROOT))
            self.assertEqual(provenance["options"]["defaults"]["reportLoci"], lab.REPORT_LOCI)
            self.assertEqual(provenance["options"]["defaults"]["fullResultSheets"], lab.FULL_RESULT_SHEETS)
            self.assertEqual(provenance["exitStatus"], 0)
            self.assertEqual(provenance["status"], "completed")
            self.assertIsNone(provenance["stderr"])
            self.assertIsInstance(provenance["wallTimeSeconds"], float)
            for key in ["python", "platform", "executable", "condaPrefix", "container"]:
                self.assertIn(key, provenance["runtimeIdentity"])

            expected_inputs = {
                str(truth_path.resolve()): truth_path,
                str(parsed_path.resolve()): parsed_path,
            }
            self.assertEqual(len(provenance["inputs"]), 2)
            self.assertEqual({record["path"] for record in provenance["inputs"]}, set(expected_inputs))
            for record in provenance["inputs"]:
                input_path = expected_inputs[record["path"]]
                self.assertEqual(record["role"], "input")
                self.assertEqual(record["sha256"], lab.sha256_file(input_path))
                self.assertEqual(record["sizeBytes"], input_path.stat().st_size)

            expected_outputs = {
                str(score_path.resolve()): score_path,
                str(report_path.resolve()): report_path,
            }
            self.assertEqual(len(provenance["outputs"]), 2)
            self.assertEqual({record["path"] for record in provenance["outputs"]}, set(expected_outputs))
            for record in provenance["outputs"]:
                output_path = expected_outputs[record["path"]]
                self.assertEqual(record["role"], "output")
                self.assertEqual(record["sha256"], lab.sha256_file(output_path))
                self.assertEqual(record["sizeBytes"], output_path.stat().st_size)

    def test_later_full_result_sheet_supersedes_duplicate_sample(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self.make_synthetic_snprc_workbook(Path(temp))
            wb = load_workbook(workbook)
            ws = wb["Full Sequencing Results 2"]
            ws.append(["Client ID", None, None, "99999", "88888"])
            ws.append(["GS ID", "Total", "Average", "LC1730", "LC1766b"])
            ws.append(["Mapped Read Count", None, None, 900, 700])
            self.append_truth_rows(
                ws,
                [
                    ("S2A1", "S2A2"),
                    ("B_A1", "B_A2"),
                ],
            )
            ws.append(["01_Mamu-A1_999g", 100, 1, 333, None])
            ws.append(["01_Mamu-A1_1766bg", 100, 1, None, 222])
            wb.save(workbook)

            extracted = lab.extract_workbook(workbook)

        self.assertEqual([sample["gs_id"] for sample in extracted["samples"]], ["LC1729", "LC1730", "LC1766b"])
        retained_lc1730 = next(sample for sample in extracted["samples"] if sample["gs_id"] == "LC1730")
        self.assertEqual(retained_lc1730["sheet"], "Full Sequencing Results 2")
        self.assertEqual(retained_lc1730["client_id"], "99999")
        self.assertEqual(retained_lc1730["mapped_reads"], 900)
        self.assertEqual(extracted["truth_calls"]["LC1730"]["MHC-A"], ["S2A1", "S2A2"])
        self.assertEqual(extracted["truth_calls"]["LC1766b"]["MHC-A"], ["B_A1", "B_A2"])

        observations = extracted["prompt_input"]["observations"]
        lc1730_observations = [row for row in observations if row["sample_id"] == "LC1730"]
        self.assertEqual({row["sheet"] for row in lc1730_observations}, {"Full Sequencing Results 2"})
        self.assertEqual([row["genotype"] for row in lc1730_observations], ["01_Mamu-A1_999g"])
        self.assertIn(
            {
                "sample_id": "LC1766b",
                "client_id": "88888",
                "report_locus": "MHC-A",
                "source_locus": "Mamu-A1",
                "genotype": "01_Mamu-A1_1766bg",
                "reads": 222,
                "sheet": "Full Sequencing Results 2",
                "row": 19,
                "sample_mapped_reads": 700,
                "read_fraction": 0.317143,
            },
            observations,
        )

        instructions = extracted["prompt_input"]["instructions"]
        self.assertIn("duplicate_sample_policy", instructions)
        superseded = extracted["prompt_input"]["superseded_samples"]
        self.assertEqual(
            superseded,
            [
                {
                    "gs_id": "LC1730",
                    "superseded_sheet": "Full Sequencing Results 1",
                    "retained_sheet": "Full Sequencing Results 2",
                    "reason": "later full-result sheet supersedes earlier occurrence for same gs_id",
                }
            ],
        )
        self.assertEqual(set(superseded[0]), {"gs_id", "superseded_sheet", "retained_sheet", "reason"})
        self.assertNotIn("A008.01", json.dumps(extracted["prompt_input"], sort_keys=True))

    def test_extract_rejects_observed_sample_without_truth(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self.make_synthetic_snprc_workbook(Path(temp))
            wb = load_workbook(workbook)
            ws = wb["Full Sequencing Results 2"]
            ws.append(["Client ID", None, None, "55555"])
            ws.append(["GS ID", "Total", "Average", "LC999"])
            ws.append(["Mapped Read Count", None, None, 500])
            ws.append(["01_Mamu-A1_999g", 100, 1, 44])
            wb.save(workbook)

            with self.assertRaisesRegex(ValueError, "missing truth.*LC999"):
                lab.extract_workbook(workbook)

    def test_extract_rejects_workbook_missing_expected_result_sheet(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "missing_sheet.xlsx"
            wb = Workbook()
            wb.active.title = "Full Sequencing Results 1"
            wb.save(path)

            with self.assertRaisesRegex(ValueError, "missing.*Full Sequencing Results 2"):
                lab.extract_workbook(path)

    def test_extract_rejects_incomplete_truth_slots(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self.make_synthetic_snprc_workbook(Path(temp))
            wb = load_workbook(workbook)
            wb["Full Sequencing Results 1"]["D6"] = None
            wb.save(workbook)

            with self.assertRaisesRegex(ValueError, "LC1729.*MHC-A"):
                lab.extract_workbook(workbook)

    def test_extract_cli_writes_artifacts_and_complete_provenance(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self.make_synthetic_snprc_workbook(root)
            output_dir = root / "out"
            argv = ["extract", "--workbook", str(workbook), "--output-dir", str(output_dir)]

            self.assertEqual(lab.main(argv), 0)

            artifact_names = ["prompt_input.json", "truth_calls.json", "samples.json", "extract.provenance.json"]
            for name in artifact_names:
                self.assertTrue((output_dir / name).is_file(), name)

            provenance = json.loads((output_dir / "extract.provenance.json").read_text())
            expected_argv = [str(Path(lab.__file__)), *argv]
            self.assertEqual(provenance["schemaVersion"], 1)
            self.assertEqual(provenance["workflowName"], "extract")
            self.assertEqual(provenance["toolName"], lab.TOOL_NAME)
            self.assertEqual(provenance["toolVersion"], lab.TOOL_VERSION)
            self.assertEqual(provenance["argv"], expected_argv)
            self.assertEqual(
                provenance["reproducibleShellCommand"],
                " ".join(shlex.quote(part) for part in [sys.executable, *expected_argv]),
            )
            self.assertEqual(provenance["options"]["workbook"], str(workbook.resolve()))
            self.assertEqual(provenance["options"]["outputDir"], str(output_dir.resolve()))
            self.assertEqual(provenance["options"]["defaults"]["workbook"], str(lab.DEFAULT_SNPRC_WORKBOOK))
            self.assertEqual(provenance["options"]["defaults"]["prompt"], str(lab.DEFAULT_PROMPT))
            self.assertEqual(provenance["options"]["defaults"]["outputRoot"], str(lab.DEFAULT_OUTPUT_ROOT))
            self.assertEqual(provenance["options"]["defaults"]["reportLoci"], lab.REPORT_LOCI)
            self.assertEqual(provenance["options"]["defaults"]["fullResultSheets"], lab.FULL_RESULT_SHEETS)
            self.assertEqual(provenance["exitStatus"], 0)
            self.assertEqual(provenance["status"], "completed")
            self.assertIsNone(provenance["stderr"])
            self.assertIsInstance(provenance["wallTimeSeconds"], float)
            for key in ["python", "platform", "executable", "condaPrefix", "container"]:
                self.assertIn(key, provenance["runtimeIdentity"])
            self.assertTrue(provenance["runtimeIdentity"]["python"])
            self.assertTrue(provenance["runtimeIdentity"]["platform"])
            self.assertEqual(provenance["runtimeIdentity"]["executable"], sys.executable)

            self.assertEqual(len(provenance["inputs"]), 1)
            input_record = provenance["inputs"][0]
            self.assertEqual(input_record["role"], "input")
            self.assertEqual(input_record["path"], str(workbook.resolve()))
            self.assertEqual(input_record["sha256"], lab.sha256_file(workbook))
            self.assertEqual(input_record["sizeBytes"], workbook.stat().st_size)

            expected_outputs = {
                str((output_dir / "prompt_input.json").resolve()): output_dir / "prompt_input.json",
                str((output_dir / "truth_calls.json").resolve()): output_dir / "truth_calls.json",
                str((output_dir / "samples.json").resolve()): output_dir / "samples.json",
            }
            self.assertEqual({record["path"] for record in provenance["outputs"]}, set(expected_outputs))
            for record in provenance["outputs"]:
                output_path = expected_outputs[record["path"]]
                self.assertEqual(record["role"], "output")
                self.assertEqual(record["sha256"], lab.sha256_file(output_path))
                self.assertEqual(record["sizeBytes"], output_path.stat().st_size)

    def test_render_prompt_cli_writes_prompt_and_complete_provenance(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            prompt_input = {
                "schema_version": 1,
                "dataset": "synthetic.xlsx",
                "instructions": {"truth_blinded": True, "report_loci": lab.REPORT_LOCI},
                "samples": [{"gs_id": "LC1729", "client_id": "44470"}],
                "observations": [
                    {
                        "sample_id": "LC1729",
                        "report_locus": "MHC-A",
                        "genotype": "01_Mamu-A1_002g",
                        "reads": 90,
                    }
                ],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            prompt_path = root / "prompt.md"
            prompt_path.write_text("# Prompt\n\n{{PROMPT_INPUT_JSON}}\n", encoding="utf-8")
            argv = ["render-prompt", "--iteration-dir", str(iteration_dir), "--prompt", str(prompt_path)]

            self.assertEqual(lab.main(argv), 0)

            output_path = iteration_dir / "rendered_prompt.md"
            provenance_path = iteration_dir / "render-prompt.provenance.json"
            self.assertTrue(output_path.is_file())
            self.assertTrue(provenance_path.is_file())
            rendered = output_path.read_text(encoding="utf-8")
            self.assertIn('"truth_blinded": true', rendered)
            self.assertIn("01_Mamu-A1_002g", rendered)

            provenance = json.loads(provenance_path.read_text())
            expected_argv = [str(Path(lab.__file__)), *argv]
            self.assertEqual(provenance["schemaVersion"], 1)
            self.assertEqual(provenance["workflowName"], "render-prompt")
            self.assertEqual(provenance["toolName"], lab.TOOL_NAME)
            self.assertEqual(provenance["toolVersion"], lab.TOOL_VERSION)
            self.assertEqual(provenance["argv"], expected_argv)
            self.assertEqual(
                provenance["reproducibleShellCommand"],
                " ".join(shlex.quote(part) for part in [sys.executable, *expected_argv]),
            )
            self.assertEqual(provenance["options"]["iterationDir"], str(iteration_dir.resolve()))
            self.assertEqual(provenance["options"]["prompt"], str(prompt_path.resolve()))
            self.assertEqual(provenance["options"]["defaults"]["workbook"], str(lab.DEFAULT_SNPRC_WORKBOOK))
            self.assertEqual(provenance["options"]["defaults"]["prompt"], str(lab.DEFAULT_PROMPT))
            self.assertEqual(provenance["options"]["defaults"]["outputRoot"], str(lab.DEFAULT_OUTPUT_ROOT))
            self.assertEqual(provenance["options"]["defaults"]["reportLoci"], lab.REPORT_LOCI)
            self.assertEqual(provenance["options"]["defaults"]["fullResultSheets"], lab.FULL_RESULT_SHEETS)
            self.assertEqual(provenance["exitStatus"], 0)
            self.assertEqual(provenance["status"], "completed")
            self.assertIsNone(provenance["stderr"])
            self.assertIsInstance(provenance["wallTimeSeconds"], float)
            for key in ["python", "platform", "executable", "condaPrefix", "container"]:
                self.assertIn(key, provenance["runtimeIdentity"])

            expected_inputs = {
                str(prompt_path.resolve()): prompt_path,
                str(prompt_input_path.resolve()): prompt_input_path,
            }
            self.assertEqual({record["path"] for record in provenance["inputs"]}, set(expected_inputs))
            for record in provenance["inputs"]:
                input_path = expected_inputs[record["path"]]
                self.assertEqual(record["role"], "input")
                self.assertEqual(record["sha256"], lab.sha256_file(input_path))
                self.assertEqual(record["sizeBytes"], input_path.stat().st_size)

            self.assertEqual(len(provenance["outputs"]), 1)
            output_record = provenance["outputs"][0]
            self.assertEqual(output_record["role"], "output")
            self.assertEqual(output_record["path"], str(output_path.resolve()))
            self.assertEqual(output_record["sha256"], lab.sha256_file(output_path))
            self.assertEqual(output_record["sizeBytes"], output_path.stat().st_size)

    def test_import_output_cli_writes_parsed_output_validation_and_complete_provenance(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            model_output = {
                "schema_version": 1,
                "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                "haplotype_definitions": [
                    {
                        "locus": "MHC-A",
                        "label": "A-A1*002-H01",
                        "supporting_genotypes": ["01_Mamu-A1_002g"],
                        "seed_samples": ["LC1729"],
                        "confidence": "high",
                        "rationale": "seed",
                    }
                ],
                "sample_calls": [
                    {
                        "sample_id": "LC1729",
                        "locus": "MHC-A",
                        "h1": "A-A1*002-H01",
                        "h2": "A-A1*002-H01",
                        "status": "called",
                        "h1_supporting_genotypes": ["01_Mamu-A1_002g"],
                        "h2_supporting_genotypes": ["01_Mamu-A1_002g"],
                        "rationale": "called",
                    }
                ],
                "unresolved": [],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text(json.dumps(model_output), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            self.assertEqual(lab.main(argv), 0)

            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            self.assertTrue(validation_path.is_file())
            self.assertTrue(parsed_path.is_file())
            self.assertTrue(provenance_path.is_file())
            self.assertEqual(json.loads(parsed_path.read_text(encoding="utf-8")), model_output)
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertTrue(validation["accepted"])
            self.assertEqual(validation["errors"], [])

            provenance = json.loads(provenance_path.read_text())
            expected_argv = [str(Path(lab.__file__)), *argv]
            self.assertEqual(provenance["schemaVersion"], 1)
            self.assertEqual(provenance["workflowName"], "import-output")
            self.assertEqual(provenance["toolName"], lab.TOOL_NAME)
            self.assertEqual(provenance["toolVersion"], lab.TOOL_VERSION)
            self.assertEqual(provenance["argv"], expected_argv)
            self.assertEqual(
                provenance["reproducibleShellCommand"],
                " ".join(shlex.quote(part) for part in [sys.executable, *expected_argv]),
            )
            self.assertEqual(provenance["options"]["iterationDir"], str(iteration_dir.resolve()))
            self.assertEqual(provenance["options"]["modelOutput"], str(model_output_path.resolve()))
            self.assertEqual(provenance["options"]["defaults"]["workbook"], str(lab.DEFAULT_SNPRC_WORKBOOK))
            self.assertEqual(provenance["options"]["defaults"]["prompt"], str(lab.DEFAULT_PROMPT))
            self.assertEqual(provenance["options"]["defaults"]["outputRoot"], str(lab.DEFAULT_OUTPUT_ROOT))
            self.assertEqual(provenance["options"]["defaults"]["reportLoci"], lab.REPORT_LOCI)
            self.assertEqual(provenance["options"]["defaults"]["fullResultSheets"], lab.FULL_RESULT_SHEETS)
            self.assertEqual(provenance["exitStatus"], 0)
            self.assertEqual(provenance["status"], "completed")
            self.assertIsNone(provenance["stderr"])
            self.assertIsInstance(provenance["wallTimeSeconds"], float)
            for key in ["python", "platform", "executable", "condaPrefix", "container"]:
                self.assertIn(key, provenance["runtimeIdentity"])

            expected_inputs = {
                str(prompt_input_path.resolve()): prompt_input_path,
                str(model_output_path.resolve()): model_output_path,
            }
            self.assertEqual(len(provenance["inputs"]), 2)
            self.assertEqual({record["path"] for record in provenance["inputs"]}, set(expected_inputs))
            for record in provenance["inputs"]:
                input_path = expected_inputs[record["path"]]
                self.assertEqual(record["role"], "input")
                self.assertEqual(record["sha256"], lab.sha256_file(input_path))
                self.assertEqual(record["sizeBytes"], input_path.stat().st_size)

            expected_outputs = {
                str(validation_path.resolve()): validation_path,
                str(parsed_path.resolve()): parsed_path,
            }
            self.assertEqual(len(provenance["outputs"]), 2)
            self.assertEqual({record["path"] for record in provenance["outputs"]}, set(expected_outputs))
            for record in provenance["outputs"]:
                output_path = expected_outputs[record["path"]]
                self.assertEqual(record["role"], "output")
                self.assertEqual(record["sha256"], lab.sha256_file(output_path))
                self.assertEqual(record["sizeBytes"], output_path.stat().st_size)

    def test_import_output_cli_records_failed_validation_provenance_without_parsed_output(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            model_output = {
                "schema_version": 1,
                "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                "haplotype_definitions": [],
                "sample_calls": [
                    {
                        "sample_id": "LC9999",
                        "locus": "MHC-Z",
                        "h1": "?",
                        "h2": "?",
                        "status": "called",
                        "h1_supporting_genotypes": ["missing"],
                        "h2_supporting_genotypes": [],
                        "rationale": "bad",
                    }
                ],
                "unresolved": [],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text(json.dumps(model_output), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            with self.assertRaisesRegex(SystemExit, "model output validation failed"):
                lab.main(argv)

            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            self.assertTrue(validation_path.is_file())
            self.assertFalse(parsed_path.exists())
            self.assertTrue(provenance_path.is_file())
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertFalse(validation["accepted"])
            self.assertTrue(any("unknown sample" in error for error in validation["errors"]))
            self.assertTrue(any("unknown locus" in error for error in validation["errors"]))
            self.assertTrue(any("unknown genotype" in error for error in validation["errors"]))

            provenance = json.loads(provenance_path.read_text())
            self.assertEqual(provenance["workflowName"], "import-output")
            self.assertEqual(provenance["exitStatus"], 1)
            self.assertEqual(provenance["status"], "failed")
            self.assertIsNotNone(provenance["stderr"])
            self.assertIn("model output validation failed", provenance["stderr"])
            self.assertEqual(len(provenance["inputs"]), 2)
            self.assertEqual(
                {record["path"] for record in provenance["inputs"]},
                {str(prompt_input_path.resolve()), str(model_output_path.resolve())},
            )
            self.assertEqual(len(provenance["outputs"]), 1)
            self.assertEqual({record["path"] for record in provenance["outputs"]}, {str(validation_path.resolve())})

    def test_import_output_cli_replaces_stale_success_artifacts_after_json_parse_failure(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text('{"schema_version": ', encoding="utf-8")
            parsed_path.write_text(json.dumps({"stale": True}), encoding="utf-8")
            validation_path.write_text(json.dumps({"schemaVersion": 1, "accepted": True, "errors": []}), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            with self.assertRaisesRegex(SystemExit, "read/parse failure"):
                lab.main(argv)

            self.assertFalse(parsed_path.exists())
            self.assertTrue(validation_path.is_file())
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertFalse(validation["accepted"])
            self.assertTrue(any("read/parse failure" in error for error in validation["errors"]))

            self.assertTrue(provenance_path.is_file())
            provenance = json.loads(provenance_path.read_text())
            self.assertEqual(provenance["workflowName"], "import-output")
            self.assertEqual(provenance["exitStatus"], 1)
            self.assertEqual(provenance["status"], "failed")
            self.assertIsNotNone(provenance["stderr"])
            self.assertIn("read/parse failure", provenance["stderr"])
            self.assertEqual({record["path"] for record in provenance["outputs"]}, {str(validation_path.resolve())})
            self.assertNotIn(str(parsed_path.resolve()), {record["path"] for record in provenance["outputs"]})

    def test_import_output_cli_rejects_extra_top_level_keys_without_parsed_output(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            model_output = {
                "schema_version": 1,
                "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                "haplotype_definitions": [],
                "sample_calls": [],
                "unresolved": [],
                "notes": "not part of the exact schema",
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text(json.dumps(model_output), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            with self.assertRaisesRegex(SystemExit, "model output validation failed"):
                lab.main(argv)

            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            self.assertTrue(validation_path.is_file())
            self.assertFalse(parsed_path.exists())
            self.assertTrue(provenance_path.is_file())
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertFalse(validation["accepted"])
            self.assertTrue(any("extra top-level key" in error and "notes" in error for error in validation["errors"]))
            provenance = json.loads(provenance_path.read_text())
            self.assertEqual(provenance["exitStatus"], 1)
            self.assertEqual(provenance["status"], "failed")
            self.assertIn("model output validation failed", provenance["stderr"])
            self.assertEqual({record["path"] for record in provenance["outputs"]}, {str(validation_path.resolve())})

    def test_import_output_cli_records_failed_provenance_for_malformed_field_types(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            model_output = {
                "schema_version": 1,
                "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                "haplotype_definitions": [
                    {
                        "locus": "MHC-A",
                        "label": "A-A1*002-H01",
                        "supporting_genotypes": ["01_Mamu-A1_002g"],
                        "seed_samples": ["LC1729"],
                        "confidence": [],
                        "rationale": "seed",
                    }
                ],
                "sample_calls": [],
                "unresolved": [],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text(json.dumps(model_output), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            with self.assertRaisesRegex(SystemExit, "model output validation failed"):
                lab.main(argv)

            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            self.assertTrue(validation_path.is_file())
            self.assertFalse(parsed_path.exists())
            self.assertTrue(provenance_path.is_file())
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertFalse(validation["accepted"])
            self.assertTrue(any("confidence" in error for error in validation["errors"]))
            provenance = json.loads(provenance_path.read_text())
            self.assertEqual(provenance["exitStatus"], 1)
            self.assertEqual(provenance["status"], "failed")
            self.assertIn("model output validation failed", provenance["stderr"])
            self.assertEqual({record["path"] for record in provenance["outputs"]}, {str(validation_path.resolve())})

    def test_import_output_cli_records_failed_provenance_when_validator_raises_unexpectedly(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            iteration_dir = root / "iteration-001"
            iteration_dir.mkdir()
            prompt_input_path = iteration_dir / "prompt_input.json"
            model_output_path = root / "model_output.json"
            prompt_input = {
                "samples": [{"gs_id": "LC1729"}],
                "instructions": {"report_loci": ["MHC-A"]},
                "observations": [
                    {"sample_id": "LC1729", "report_locus": "MHC-A", "genotype": "01_Mamu-A1_002g", "reads": 90}
                ],
            }
            model_output = {
                "schema_version": 1,
                "prompt_version": "generalist_macaque_mhc_haplotyping_v1",
                "haplotype_definitions": [],
                "sample_calls": [],
                "unresolved": [],
            }
            prompt_input_path.write_text(json.dumps(prompt_input), encoding="utf-8")
            model_output_path.write_text(json.dumps(model_output), encoding="utf-8")
            argv = ["import-output", "--iteration-dir", str(iteration_dir), "--model-output", str(model_output_path)]

            original_validate_model_output = lab.validate_model_output

            def raise_unexpected_validation_error(output, prompt_input):
                raise RuntimeError("boom")

            lab.validate_model_output = raise_unexpected_validation_error
            try:
                with self.assertRaisesRegex(SystemExit, "model output validation failed"):
                    lab.main(argv)
            finally:
                lab.validate_model_output = original_validate_model_output

            validation_path = iteration_dir / "model_output_validation.json"
            parsed_path = iteration_dir / "parsed_model_output.json"
            provenance_path = iteration_dir / "import-output.provenance.json"
            self.assertTrue(validation_path.is_file())
            self.assertFalse(parsed_path.exists())
            self.assertTrue(provenance_path.is_file())
            validation = json.loads(validation_path.read_text(encoding="utf-8"))
            self.assertFalse(validation["accepted"])
            self.assertEqual(validation["errors"], ["model output validation raised unexpected error: boom"])
            provenance = json.loads(provenance_path.read_text())
            self.assertEqual(provenance["exitStatus"], 1)
            self.assertEqual(provenance["status"], "failed")
            self.assertIn("model output validation raised unexpected error: boom", provenance["stderr"])
            self.assertEqual({record["path"] for record in provenance["outputs"]}, {str(validation_path.resolve())})


if __name__ == "__main__":
    unittest.main()
